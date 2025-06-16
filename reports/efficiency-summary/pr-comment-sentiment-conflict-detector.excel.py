#!/usr/bin/env python3
"""
PR Comment Sentiment & Conflict Detector - Excel Report Generator

This module creates comprehensive Excel reports with multiple sheets,
interactive charts, and formulas for analyzing PR data and ML insights.
"""

import json
import argparse
import pandas as pd
import numpy as np
from pathlib import Path
import logging
import traceback
from datetime import datetime
from typing import Dict, List, Any, Optional, Union
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.formatting import Rule
from rich.console import Console
from rich.markup import escape

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Initialize Rich console
console = Console()

def safe_console_print(message: str, style: str = "") -> None:
    """
    Safely print messages to console with Rich markup, handling errors gracefully.
    
    Args:
        message: The message to print
        style: Rich style to apply (e.g., "red", "green", "bold")
    """
    escaped = escape(str(message))
    try:
        if style:
            console.print(f"[{style}]{escaped}[/{style}]")
        else:
            console.print(escaped)
    except Exception as e:
        console.print(f"[red]⚠ Print error:[/red] {escape(str(e))}")
        console.print(escaped)


class ExcelReportGenerator:
    """Generates comprehensive Excel reports for PR analysis"""
    
    def __init__(self, output_path: Union[str, Path]):
        """
        Initialize the Excel report generator.
        
        Args:
            output_path: Path where the Excel file will be saved
        """
        self.output_path = Path(output_path)
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        
        # Define styles
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def create_comprehensive_report(self, data: Dict[str, Any], ml_insights: Optional[Dict[str, Any]] = None) -> None:
        """
        Create comprehensive Excel report from PR analysis data.
        
        Args:
            data: PR analysis data dictionary
            ml_insights: Optional ML insights data dictionary
        """
        logger.info("Creating comprehensive Excel report...")
        safe_console_print("📊 Generating Excel report...", "blue")
        
        try:
            # 1. Summary Dashboard
            self._create_summary_sheet(data, ml_insights)
            
            # 2. PR Details
            self._create_pr_details_sheet(data)
            
            # 3. Sentiment Analysis
            self._create_sentiment_sheet(data)
            
            # 4. Conflict Analysis
            self._create_conflict_sheet(data)
            
            # 5. Contributor Metrics
            self._create_contributor_sheet(data)
            
            # 6. Trends Analysis
            self._create_trends_sheet(data)
            
            # 7. ML Insights (if available)
            if ml_insights:
                self._create_ml_insights_sheet(ml_insights)
            
            # 8. Formulas Reference
            self._create_formulas_sheet(data)
            
            # Save workbook
            self.workbook.save(self.output_path)
            logger.info(f"Excel report saved to {self.output_path}")
            safe_console_print(f"✅ Excel report saved to {self.output_path}", "green")
            
        except Exception as e:
            logger.error(f"Failed to create comprehensive report: {e}")
            safe_console_print(f"Error creating report: {e}", "red")
            raise
        
    def _create_summary_sheet(self, data: Dict[str, Any], ml_insights: Optional[Dict[str, Any]] = None) -> None:
        """Create summary dashboard sheet"""
        try:
            ws = self.workbook.create_sheet("Summary Dashboard")
            
            # Title
            ws['A1'] = "PR Comment Sentiment & Conflict Detection - Summary Dashboard"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:F1')
            
            # Metadata
            metadata = data.get('metadata', {})
            date_range = data.get('date_range', {})
            
            ws['A3'] = "Report Information"
            ws['A3'].font = self.title_font
            
            info_data = [
                ['Generated At', metadata.get('generated_at', '')],
                ['Analysis Period', f"{date_range.get('start_date', '')} to {date_range.get('end_date', '')}"],
                ['Target', metadata.get('target', '')],
                ['Repository', metadata.get('repository', 'N/A')],
                ['Duration (Days)', date_range.get('duration_days', '')]
            ]
            
            for i, (key, value) in enumerate(info_data, 4):
                ws[f'A{i}'] = key
                ws[f'B{i}'] = value
                ws[f'A{i}'].font = Font(bold=True)
            
            # Key Metrics
            ws['D3'] = "Key Metrics"
            ws['D3'].font = self.title_font
            
            summary = data.get('summary', {})
            metrics_data = [
                ['Total PRs', summary.get('total_prs', 0)],
                ['Merged PRs', summary.get('merged_prs', 0)],
                ['Merge Rate (%)', summary.get('merge_rate', 0)],
                ['Total Comments', summary.get('total_comments', 0)],
                ['Total Conflicts', summary.get('total_conflicts_detected', 0)],
                ['Conflict Rate (%)', summary.get('conflict_rate', 0)],
                ['Avg Comments/PR', summary.get('average_comments_per_pr', 0)],
                ['High Severity Conflicts', summary.get('high_severity_conflicts', 0)]
            ]
            
            for i, (key, value) in enumerate(metrics_data, 4):
                ws[f'D{i}'] = key
                ws[f'E{i}'] = value
                ws[f'D{i}'].font = Font(bold=True)
                
                # Add formulas for calculated metrics
                if 'Rate' in key:
                    ws[f'F{i}'] = f"=E{i}/100"
                    ws[f'F{i}'].number_format = '0.00%'
            
            # Sentiment Overview
            sentiment_data = data.get('sentiment_analysis', {})
            ws['A12'] = "Sentiment Analysis"
            ws['A12'].font = self.title_font
            
            sentiment_info = [
                ['Overall Sentiment Score', sentiment_data.get('overall_sentiment_score', 0)],
                ['Overall Sentiment Label', sentiment_data.get('overall_sentiment_label', 'neutral')],
                ['Positive PRs', sentiment_data.get('sentiment_distribution', {}).get('positive', 0)],
                ['Neutral PRs', sentiment_data.get('sentiment_distribution', {}).get('neutral', 0)],
                ['Negative PRs', sentiment_data.get('sentiment_distribution', {}).get('negative', 0)]
            ]
            
            for i, (key, value) in enumerate(sentiment_info, 13):
                ws[f'A{i}'] = key
                ws[f'B{i}'] = value
                ws[f'A{i}'].font = Font(bold=True)
            
            # ML Insights Summary (if available)
            if ml_insights:
                ml_summary = ml_insights.get('insights', {}).get('summary', {})
                ws['D12'] = "ML Analysis Summary"
                ws['D12'].font = self.title_font
                
                ml_data = [
                    ['PRs Analyzed (ML)', ml_summary.get('total_prs_analyzed', 0)],
                    ['Avg Sentiment (ML)', ml_summary.get('avg_sentiment_score', 0)],
                    ['Conflict Rate (ML)', ml_summary.get('conflict_rate', 0)],
                    ['Merge Rate (ML)', ml_summary.get('merge_rate', 0)]
                ]
                
                for i, (key, value) in enumerate(ml_data, 13):
                    ws[f'D{i}'] = key
                    ws[f'E{i}'] = value
                    ws[f'D{i}'].font = Font(bold=True)
            
            # Apply formatting
            self._apply_table_formatting(ws, 'A3:B8')
            self._apply_table_formatting(ws, 'D3:F11')
            self._apply_table_formatting(ws, 'A12:B17')
            if ml_insights:
                self._apply_table_formatting(ws, 'D12:E16')
            
            # Auto-adjust column widths with error handling
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating summary sheet: {e}")
            safe_console_print(f"Error creating summary sheet: {e}", "red")
            raise
    
    def _create_pr_details_sheet(self, data: Dict[str, Any]) -> None:
        """Create detailed PR information sheet"""
        try:
            ws = self.workbook.create_sheet("PR Details")
            
            # Get PR data
            pr_data = data.get('detailed_analysis', {}).get('pull_requests', [])
            
            if not pr_data:
                ws['A1'] = "No PR data available"
                return
            
            # Create DataFrame
            df = pd.DataFrame(pr_data)
            
            # Add calculated columns with formulas
            df['conflict_ratio'] = df['conflict_count'] / (df['total_comments'] + 1)  # +1 to avoid division by zero
            df['engagement_score'] = df['total_comments'] + df['total_reviews']
            
            # Convert DataFrame to Excel
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Apply header formatting
            for cell in ws[1]:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
            
            # Add formulas in empty columns
            last_col = ws.max_column
            
            # Add status indicator formula
            ws.cell(1, last_col + 1, "Status Indicator")
            status_col = get_column_letter(last_col + 1)
            
            for row in range(2, ws.max_row + 1):
                # Status based on state and conflicts
                if 'state' in df.columns and 'conflict_count' in df.columns:
                    state_col = get_column_letter(df.columns.get_loc("state") + 1)
                    conflict_col = get_column_letter(df.columns.get_loc("conflict_count") + 1)
                    formula = f'=IF(AND(ISTEXT({state_col}{row}), {state_col}{row}="merged"), "✅ Merged", IF({conflict_col}{row}>2, "⚠️ High Conflict", "📝 Normal"))'
                    ws[f'{status_col}{row}'] = formula
            
            # Apply conditional formatting
            self._apply_conditional_formatting(ws, df)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating PR details sheet: {e}")
            safe_console_print(f"Error creating PR details sheet: {e}", "red")
            raise
    
    def _create_sentiment_sheet(self, data: Dict[str, Any]) -> None:
        """Create sentiment analysis sheet"""
        try:
            ws = self.workbook.create_sheet("Sentiment Analysis")
            
            sentiment_data = data.get('sentiment_analysis', {})
            pr_sentiments = sentiment_data.get('pr_sentiments', [])
            
            if not pr_sentiments:
                ws['A1'] = "No sentiment data available"
                return
            
            # Create DataFrame
            df = pd.DataFrame(pr_sentiments)
            if 'created_at' in df.columns:
                df['created_date'] = pd.to_datetime(df['created_at']).dt.date
            
            # Convert DataFrame to Excel
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Apply header formatting
            for cell in ws[1]:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
            
            # Add summary statistics
            ws['H1'] = "Sentiment Statistics"
            ws['H1'].font = self.title_font
            
            if 'sentiment_score' in df.columns:
                sentiment_col = get_column_letter(df.columns.get_loc('sentiment_score') + 1)
                
                stats = [
                    ['Average Sentiment', f'=AVERAGE({sentiment_col}2:{sentiment_col}{ws.max_row})'],
                    ['Median Sentiment', f'=MEDIAN({sentiment_col}2:{sentiment_col}{ws.max_row})'],
                    ['Standard Deviation', f'=STDEV({sentiment_col}2:{sentiment_col}{ws.max_row})'],
                    ['Positive Count', f'=COUNTIF({sentiment_col}2:{sentiment_col}{ws.max_row},">0.05")'],
                    ['Negative Count', f'=COUNTIF({sentiment_col}2:{sentiment_col}{ws.max_row},"<-0.05")'],
                    ['Neutral Count', f'=COUNTIFS({sentiment_col}2:{sentiment_col}{ws.max_row},"<=0.05",{sentiment_col}2:{sentiment_col}{ws.max_row},">=-0.05")']
                ]
                
                for i, (label, formula) in enumerate(stats, 2):
                    ws[f'H{i}'] = label
                    ws[f'I{i}'] = formula
                    ws[f'H{i}'].font = Font(bold=True)
                
                # Add conditional formatting for sentiment scores
                sentiment_range = f'{sentiment_col}2:{sentiment_col}{ws.max_row}'
                
                # Green for positive sentiment
                ws.conditional_formatting.add(sentiment_range,
                    CellIsRule(operator='greaterThan', formula=['0.05'],
                              fill=PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')))
                
                # Red for negative sentiment
                ws.conditional_formatting.add(sentiment_range,
                    CellIsRule(operator='lessThan', formula=['-0.05'],
                              fill=PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')))
            
            # Create sentiment trend chart
            self._create_sentiment_chart(ws, df)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating sentiment sheet: {e}")
            safe_console_print(f"Error creating sentiment sheet: {e}", "red")
            raise
    
    def _create_conflict_sheet(self, data: Dict[str, Any]) -> None:
        """Create conflict analysis sheet"""
        try:
            ws = self.workbook.create_sheet("Conflict Analysis")
            
            conflict_data = data.get('conflict_detection', {})
            conflicts = conflict_data.get('conflicts', [])
            
            if not conflicts:
                ws['A1'] = "No conflict data available"
                ws['A2'] = "This indicates a healthy repository with minimal disagreements"
                ws['A2'].font = Font(italic=True, color="666666")
                return
            
            # Create DataFrame
            df = pd.DataFrame(conflicts)
            
            # Convert DataFrame to Excel
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Apply header formatting
            for cell in ws[1]:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
            
            # Add conflict statistics
            ws['L1'] = "Conflict Statistics"
            ws['L1'].font = self.title_font
            
            if 'severity' in df.columns:
                severity_col = get_column_letter(df.columns.get_loc('severity') + 1)
                
                stats = [
                    ['Total Conflicts', f'=COUNTA({severity_col}2:{severity_col}{ws.max_row})'],
                    ['High Severity', f'=COUNTIF({severity_col}2:{severity_col}{ws.max_row},"high")'],
                    ['Medium Severity', f'=COUNTIF({severity_col}2:{severity_col}{ws.max_row},"medium")'],
                    ['Low Severity', f'=COUNTIF({severity_col}2:{severity_col}{ws.max_row},"low")'],
                    ['High Severity %', f'=M3/M2*100'],
                    ['Unique PRs with Conflicts', f'=SUMPRODUCT(1/COUNTIF(A2:A{ws.max_row},A2:A{ws.max_row}))']
                ]
                
                for i, (label, formula) in enumerate(stats, 2):
                    ws[f'L{i}'] = label
                    ws[f'M{i}'] = formula
                    ws[f'L{i}'].font = Font(bold=True)
                    
                    if '%' in label:
                        ws[f'M{i}'].number_format = '0.00%'
                
                # Apply conditional formatting for severity
                severity_range = f'{severity_col}2:{severity_col}{ws.max_row}'
                
                # High severity - Red
                ws.conditional_formatting.add(severity_range,
                    CellIsRule(operator='equal', formula=['"high"'],
                              fill=PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')))
                
                # Medium severity - Orange
                ws.conditional_formatting.add(severity_range,
                    CellIsRule(operator='equal', formula=['"medium"'],
                              fill=PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')))
                
                # Low severity - Yellow
                ws.conditional_formatting.add(severity_range,
                    CellIsRule(operator='equal', formula=['"low"'],
                              fill=PatternFill(start_color='6BCF7F', end_color='6BCF7F', fill_type='solid')))
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating conflict sheet: {e}")
            safe_console_print(f"Error creating conflict sheet: {e}", "red")
            raise
    
    def _create_contributor_sheet(self, data: Dict[str, Any]) -> None:
        """Create contributor metrics sheet"""
        try:
            ws = self.workbook.create_sheet("Contributors")
            
            contributor_data = data.get('contributor_metrics', {})
            contributors = contributor_data.get('contributors', [])
            
            if not contributors:
                ws['A1'] = "No contributor data available"
                return
            
            # Create DataFrame
            df = pd.DataFrame(contributors)
            
            # Add calculated columns
            df['total_activity'] = df['prs_authored'] + df['reviews_given'] + df['comments_made']
            df['activity_ratio'] = df['prs_authored'] / (df['total_activity'] + 1)  # Avoid division by zero
            df['review_ratio'] = df['reviews_given'] / (df['total_activity'] + 1)
            
            # Convert DataFrame to Excel
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Apply header formatting
            for cell in ws[1]:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
            
            # Add contributor rankings
            last_col = ws.max_column
            
            # Most Active Contributors
            ws.cell(1, last_col + 2, "Top Contributors")
            ws.cell(1, last_col + 2).font = self.title_font
            
            if 'total_activity' in df.columns:
                activity_col = get_column_letter(df.columns.get_loc('total_activity') + 1)
                
                rankings = [
                    ['Most Active', f'=INDEX(A:A,MATCH(MAX({activity_col}:{activity_col}),{activity_col}:{activity_col},0))'],
                    ['Most PRs', f'=INDEX(A:A,MATCH(MAX(B:B),B:B,0))'],
                    ['Most Reviews', f'=INDEX(A:A,MATCH(MAX(C:C),C:C,0))'],
                    ['Most Comments', f'=INDEX(A:A,MATCH(MAX(D:D),D:D,0))'],
                    ['Most Conflicts', f'=INDEX(A:A,MATCH(MAX(E:E),E:E,0))']
                ]
                
                for i, (label, formula) in enumerate(rankings, 2):
                    ws.cell(i, last_col + 2, label)
                    ws.cell(i, last_col + 3, formula)
                    ws.cell(i, last_col + 2).font = Font(bold=True)
            
            # Create contributor activity chart
            self._create_contributor_chart(ws, df)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating contributor sheet: {e}")
            safe_console_print(f"Error creating contributor sheet: {e}", "red")
            raise
    
    def _create_trends_sheet(self, data: Dict[str, Any]) -> None:
        """Create trends analysis sheet"""
        try:
            ws = self.workbook.create_sheet("Trends")
            
            trends_data = data.get('trends', {})
            
            # Weekly trends
            weekly_trends = trends_data.get('weekly_trends', [])
            if weekly_trends:
                ws['A1'] = "Weekly Trends"
                ws['A1'].font = self.title_font
                
                weekly_df = pd.DataFrame(weekly_trends)
                
                # Insert weekly data
                row_start = 3
                for r in dataframe_to_rows(weekly_df, index=False, header=True):
                    ws.append(r)
                    if ws.max_row == row_start:  # Header row
                        for cell in ws[ws.max_row]:
                            cell.font = self.header_font
                            cell.fill = self.header_fill
            
            # Monthly trends
            monthly_trends = trends_data.get('monthly_trends', [])
            if monthly_trends:
                start_row = ws.max_row + 3
                ws[f'A{start_row}'] = "Monthly Trends"
                ws[f'A{start_row}'].font = self.title_font
                
                monthly_df = pd.DataFrame(monthly_trends)
                
                # Insert monthly data
                for r in dataframe_to_rows(monthly_df, index=False, header=True):
                    ws.append(r)
                    if ws.max_row == start_row + 2:  # Header row
                        for cell in ws[ws.max_row]:
                            cell.font = self.header_font
                            cell.fill = self.header_fill
            
            # Add trend analysis formulas
            if weekly_trends:
                ws['G3'] = "Weekly Analysis"
                ws['G3'].font = self.title_font
                
                analysis_formulas = [
                    ['Peak Week (PRs)', '=INDEX(A:A,MATCH(MAX(B:B),B:B,0))'],
                    ['Avg PRs/Week', '=AVERAGE(B:B)'],
                    ['Avg Sentiment/Week', '=AVERAGE(D:D)'],
                    ['Total Conflicts', '=SUM(C:C)']
                ]
                
                for i, (label, formula) in enumerate(analysis_formulas, 4):
                    ws[f'G{i}'] = label
                    ws[f'H{i}'] = formula
                    ws[f'G{i}'].font = Font(bold=True)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating trends sheet: {e}")
            safe_console_print(f"Error creating trends sheet: {e}", "red")
            raise
    
    def _create_ml_insights_sheet(self, ml_insights: Dict[str, Any]) -> None:
        """Create ML insights sheet"""
        try:
            ws = self.workbook.create_sheet("ML Insights")
            
            insights = ml_insights.get('insights', {})
            
            # Summary insights
            ws['A1'] = "ML Analysis Summary"
            ws['A1'].font = Font(bold=True, size=14)
            
            summary = insights.get('summary', {})
            row = 3
            for key, value in summary.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Collaboration insights
            row += 2
            ws[f'A{row}'] = "Collaboration Patterns"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            collab_insights = insights.get('collaboration_insights', {})
            if collab_insights:
                # Most efficient pattern
                efficient = collab_insights.get('most_efficient_pattern', {})
                if efficient:
                    ws[f'A{row}'] = "Most Efficient Pattern"
                    ws[f'B{row}'] = efficient.get('name', '')
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1
                    
                    chars = efficient.get('characteristics', {})
                    for key, value in chars.items():
                        ws[f'B{row}'] = f"{key.replace('_', ' ').title()}: {value}"
                        row += 1
            
            # Risk insights
            row += 2
            ws[f'A{row}'] = "Risk Analysis"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            risk_insights = insights.get('risk_insights', {})
            for key, value in risk_insights.items():
                if key != 'key_risk_factors':
                    ws[f'A{row}'] = key.replace('_', ' ').title()
                    ws[f'B{row}'] = value
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1
            
            # Recommendations
            row += 2
            ws[f'A{row}'] = "Recommendations"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            recommendations = insights.get('recommendations', [])
            for rec in recommendations:
                ws[f'A{row}'] = f"[{rec.get('priority', 'Medium')}] {rec.get('category', '')}"
                ws[f'B{row}'] = rec.get('recommendation', '')
                ws[f'C{row}'] = rec.get('reasoning', '')
                
                # Color code by priority
                if rec.get('priority') == 'High':
                    ws[f'A{row}'].fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                elif rec.get('priority') == 'Medium':
                    ws[f'A{row}'].fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
                
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating ML insights sheet: {e}")
            safe_console_print(f"Error creating ML insights sheet: {e}", "red")
            raise
    
    def _create_formulas_sheet(self, data: Dict[str, Any]) -> None:
        """Create formulas reference sheet"""
        try:
            ws = self.workbook.create_sheet("Formulas Reference")
            
            ws['A1'] = "Formulas and Calculations Reference"
            ws['A1'].font = Font(bold=True, size=16)
            
            formulas = data.get('formulas', {})
            
            headers = ['Formula Name', 'Formula', 'Description', 'Example Calculation']
            for i, header in enumerate(headers, 1):
                ws.cell(3, i, header)
                ws.cell(3, i).font = self.header_font
                ws.cell(3, i).fill = self.header_fill
                ws.cell(3, i).border = self.border
            
            # Formula descriptions
            formula_descriptions = {
                'sentiment_score': 'Calculates average sentiment from individual comment scores',
                'conflict_rate': 'Percentage of PRs with detected conflicts',
                'merge_rate': 'Percentage of PRs that were successfully merged',
                'network_density': 'Ratio of actual interactions to possible interactions',
                'average_comments_per_pr': 'Mean number of comments across all PRs',
                'contributor_activity_score': 'Combined activity metric for contributors'
            }
            
            row = 4
            for name, formula in formulas.items():
                ws[f'A{row}'] = name.replace('_', ' ').title()
                ws[f'B{row}'] = formula
                ws[f'C{row}'] = formula_descriptions.get(name, 'Custom calculation')
                
                # Add example calculation
                if name == 'merge_rate':
                    ws[f'D{row}'] = '=MERGED_PRS/TOTAL_PRS*100'
                elif name == 'sentiment_score':
                    ws[f'D{row}'] = '=AVERAGE(SentimentAnalysis!C:C)'
                elif name == 'conflict_rate':
                    ws[f'D{row}'] = '=COUNTIF(ConflictAnalysis!E:E,">0")/COUNT(ConflictAnalysis!E:E)*100'
                
                row += 1
            
            # Apply table formatting
            self._apply_table_formatting(ws, f'A3:D{row-1}')
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating formulas sheet: {e}")
            safe_console_print(f"Error creating formulas sheet: {e}", "red")
            raise
    
    def _apply_table_formatting(self, ws, range_str: str) -> None:
        """Apply consistent table formatting"""
        try:
            for row in ws[range_str]:
                for cell in row:
                    cell.border = self.border
                    if cell.row == ws[range_str][0][0].row:  # Header row
                        cell.font = self.header_font
                        cell.fill = self.header_fill
        except Exception as e:
            logger.warning(f"Could not apply table formatting to {range_str}: {e}")
    
    def _apply_conditional_formatting(self, ws, df: pd.DataFrame) -> None:
        """Apply conditional formatting based on data"""
        try:
            if 'sentiment_score' in df.columns:
                sentiment_col = get_column_letter(df.columns.get_loc('sentiment_score') + 1)
                sentiment_range = f'{sentiment_col}2:{sentiment_col}{ws.max_row}'
                
                # Color scale for sentiment
                rule = ColorScaleRule(
                    start_type='num', start_value=-1, start_color='FF6B6B',
                    mid_type='num', mid_value=0, mid_color='FFFFFF',
                    end_type='num', end_value=1, end_color='90EE90'
                )
                ws.conditional_formatting.add(sentiment_range, rule)
            
            if 'conflict_count' in df.columns:
                conflict_col = get_column_letter(df.columns.get_loc('conflict_count') + 1)
                conflict_range = f'{conflict_col}2:{conflict_col}{ws.max_row}'
                
                # Highlight high conflict counts
                ws.conditional_formatting.add(conflict_range,
                    CellIsRule(operator='greaterThan', formula=['2'],
                              fill=PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')))
        except Exception as e:
            logger.warning(f"Could not apply conditional formatting: {e}")
    
    def _create_sentiment_chart(self, ws, df: pd.DataFrame) -> None:
        """Create sentiment trend chart"""
        try:
            chart = LineChart()
            chart.title = "Sentiment Trends"
            chart.style = 2
            chart.y_axis.title = 'Sentiment Score'
            chart.x_axis.title = 'PR Index'
            
            # Add data (simplified)
            if 'sentiment_score' in df.columns:
                data_range = Reference(ws, min_col=df.columns.get_loc('sentiment_score') + 1, 
                                     min_row=1, max_row=min(50, ws.max_row))
                chart.add_data(data_range, titles_from_data=True)
                
                ws.add_chart(chart, "K2")
        except Exception as e:
            logger.warning(f"Could not create sentiment chart: {e}")
    
    def _create_contributor_chart(self, ws, df: pd.DataFrame) -> None:
        """Create contributor activity chart"""
        try:
            chart = BarChart()
            chart.title = "Top Contributors by Activity"
            chart.style = 10
            chart.y_axis.title = 'Activity Count'
            chart.x_axis.title = 'Contributors'
            
            # Add data for top 10 contributors
            if 'total_activity' in df.columns:
                data_range = Reference(ws, min_col=df.columns.get_loc('total_activity') + 1, 
                                      min_row=1, max_row=min(11, ws.max_row))
                chart.add_data(data_range, titles_from_data=True)
                
                ws.add_chart(chart, "K2")
        except Exception as e:
            logger.warning(f"Could not create contributor chart: {e}")
    
    def _auto_adjust_columns(self, ws) -> None:
        """
        Auto-adjust column widths with proper error handling.
        
        This fixes the original error by providing a default value when 
        all cells in a column are empty.
        """
        try:
            for column in ws.columns:
                # Calculate max length with default value for empty columns
                max_length = max(
                    (len(str(cell.value)) for cell in column if cell.value is not None),
                    default=10  # Default width for empty columns
                )
                # Set column width with reasonable bounds
                adjusted_width = min(max(max_length + 2, 8), 50)
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        except Exception as e:
            logger.warning(f"Could not auto-adjust columns: {e}")
            # Set a default width for all columns as fallback
            for i in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(i)].width = 15


def main() -> None:
    """Main function for Excel report generation"""
    parser = argparse.ArgumentParser(
        description='PR Comment Sentiment & Conflict Detector - Excel Report Generator',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python -m pr_excel_generator --input data.json --output report.xlsx
  python -m pr_excel_generator -i analysis.json -o results.xlsx --insights ml_data.json -v
        """
    )
    parser.add_argument(
        '--input', '-i', 
        required=True, 
        help='Input JSON file path containing PR analysis data'
    )
    parser.add_argument(
        '--output', '-o', 
        default='./reports/pr_analysis.xlsx', 
        help='Output Excel file path (default: ./reports/pr_analysis.xlsx)'
    )
    parser.add_argument(
        '--insights', 
        help='ML insights JSON file path (optional)'
    )
    parser.add_argument(
        '--verbose', '-v', 
        action='store_true', 
        help='Enable verbose output'
    )
    
    args = parser.parse_args()
    
    if args.verbose:
        logger.setLevel(logging.DEBUG)
        safe_console_print("🔍 Verbose mode enabled", "blue")
    
    try:
        # Load main data
        input_path = Path(args.input)
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_path}")
        
        safe_console_print(f"📂 Loading data from {input_path}", "blue")
        logger.info(f"Loading data from {input_path}")
        
        with open(input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Validate data structure
        if not isinstance(data, dict):
            raise ValueError("Input data must be a JSON object")
        
        # Load ML insights if provided
        ml_insights = None
        if args.insights:
            insights_path = Path(args.insights)
            if not insights_path.exists():
                logger.warning(f"ML insights file not found: {insights_path}")
                safe_console_print(f"⚠️  ML insights file not found: {insights_path}", "yellow")
            else:
                safe_console_print(f"🤖 Loading ML insights from {insights_path}", "blue")
                logger.info(f"Loading ML insights from {insights_path}")
                with open(insights_path, 'r', encoding='utf-8') as f:
                    ml_insights = json.load(f)
        
        # Create output directory
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Generate Excel report
        safe_console_print("🚀 Starting Excel report generation...", "green")
        generator = ExcelReportGenerator(output_path)
        generator.create_comprehensive_report(data, ml_insights)
        
        # Print summary
        safe_console_print("\n📊 Excel Report Generated Successfully!", "green")
        safe_console_print(f"   📄 File: {output_path}", "white")
        safe_console_print("   📋 Sheets: Summary Dashboard, PR Details, Sentiment Analysis, Conflict Analysis", "white")
        safe_console_print("   📈 Charts: Sentiment trends, contributor activity", "white")
        safe_console_print("   📐 Formulas: Interactive calculations and references", "white")
        if ml_insights:
            safe_console_print(f"   🤖 ML Insights: Included from {args.insights}", "white")
        
        # Display file size
        try:
            file_size = output_path.stat().st_size
            file_size_mb = file_size / (1024 * 1024)
            safe_console_print(f"   💾 File Size: {file_size_mb:.2f} MB", "white")
        except:
            pass
        
    except FileNotFoundError as e:
        error_msg = f"File not found: {e}"
        logger.error(error_msg)
        safe_console_print(error_msg, "red")
        return 1
    except json.JSONDecodeError as e:
        error_msg = f"Invalid JSON format: {e}"
        logger.error(error_msg)
        safe_console_print(error_msg, "red")
        return 1
    except ValueError as e:
        error_msg = f"Data validation error: {e}"
        logger.error(error_msg)
        safe_console_print(error_msg, "red")
        return 1
    except Exception as e:
        error_msg = f"Excel generation failed: {e}"
        logger.error(error_msg)
        safe_console_print(error_msg, "red")
        if args.verbose:
            safe_console_print("📋 Full traceback:", "yellow")
            safe_console_print(traceback.format_exc(), "red")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())