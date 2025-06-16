#!/usr/bin/env python3
"""
Stale PR Audit Excel Report Generator
Creates interactive Excel workbooks with multiple sheets and formulas.
"""

import json
import sys
import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime
import warnings
import logging
import traceback

warnings.filterwarnings('ignore')

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

def safe_float(value, default: float = 0.0) -> float:
    """Safely convert value to float with detailed logging"""
    try:
        if value is None or value == '':
            logger.debug(f"safe_float: Converting None/empty to default {default}")
            return default
        
        # Handle boolean values
        if isinstance(value, bool):
            result = float(value)
            logger.debug(f"safe_float: Converted bool {value} to {result}")
            return result
            
        # Handle string representations of numbers
        if isinstance(value, str):
            # Remove whitespace and common non-numeric characters
            cleaned = value.strip().replace(',', '').replace('%', '')
            if cleaned == '':
                logger.debug(f"safe_float: Empty string after cleaning, using default {default}")
                return default
            result = float(cleaned)
            logger.debug(f"safe_float: Converted string '{value}' to {result}")
            return result
            
        # Direct conversion for numeric types
        result = float(value)
        logger.debug(f"safe_float: Converted {type(value).__name__} {value} to {result}")
        return result
        
    except (ValueError, TypeError) as e:
        logger.warning(f"safe_float: Failed to convert {value} (type: {type(value).__name__}) to float: {e}. Using default {default}")
        return default
    except Exception as e:
        logger.error(f"safe_float: Unexpected error converting {value}: {e}. Using default {default}")
        return default

def safe_int(value, default: int = 0) -> int:
    """Safely convert value to int with detailed logging"""
    try:
        if value is None or value == '':
            logger.debug(f"safe_int: Converting None/empty to default {default}")
            return default
        
        # Handle boolean values
        if isinstance(value, bool):
            result = int(value)
            logger.debug(f"safe_int: Converted bool {value} to {result}")
            return result
            
        # Handle string representations of numbers
        if isinstance(value, str):
            # Remove whitespace and common non-numeric characters
            cleaned = value.strip().replace(',', '')
            if cleaned == '':
                logger.debug(f"safe_int: Empty string after cleaning, using default {default}")
                return default
            # Convert through float first to handle decimal strings
            result = int(float(cleaned))
            logger.debug(f"safe_int: Converted string '{value}' to {result}")
            return result
            
        # Convert through float first to handle edge cases
        result = int(float(value))
        logger.debug(f"safe_int: Converted {type(value).__name__} {value} to {result}")
        return result
        
    except (ValueError, TypeError) as e:
        logger.warning(f"safe_int: Failed to convert {value} (type: {type(value).__name__}) to int: {e}. Using default {default}")
        return default
    except Exception as e:
        logger.error(f"safe_int: Unexpected error converting {value}: {e}. Using default {default}")
        return default

def safe_percentage(value, default: float = 0.0) -> float:
    """Safely convert value to percentage (0-1 range) with detailed logging"""
    try:
        raw_value = safe_float(value, default)
        
        # If value is already in 0-1 range, return as is
        if 0 <= raw_value <= 1:
            logger.debug(f"safe_percentage: Value {raw_value} already in percentage range")
            return raw_value
        
        # If value is in 0-100 range, convert to 0-1
        if 0 <= raw_value <= 100:
            result = raw_value / 100.0
            logger.debug(f"safe_percentage: Converted {raw_value}% to {result}")
            return result
        
        # For values outside expected ranges, log warning and use default
        logger.warning(f"safe_percentage: Value {raw_value} outside expected percentage range. Using default {default}")
        return default
        
    except Exception as e:
        logger.error(f"safe_percentage: Unexpected error converting {value}: {e}. Using default {default}")
        return default

class StalePRExcelGenerator:
    def __init__(self, verbose: bool = False):
        self.verbose = verbose
        self.wb = Workbook()
        self.data = {}
        self.ml_insights = None
        logger.info(f"Initialized StalePRExcelGenerator with verbose={verbose}")
        self.setup_styles()
    
    def setup_styles(self):
        """Define styles for the workbook"""
        logger.debug("Setting up Excel styles")
        self.styles = {
            'header': Font(bold=True, color='FFFFFF'),
            'header_fill': PatternFill(start_color='2E3440', end_color='2E3440', fill_type='solid'),
            'subheader': Font(bold=True, color='2E3440'),
            'subheader_fill': PatternFill(start_color='D8DEE9', end_color='D8DEE9', fill_type='solid'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center')
        }
        logger.debug("Excel styles configured successfully")
    
    def load_data(self, input_file: str) -> dict:
        """Load report data from JSON file"""
        logger.info(f"Loading data from {input_file}")
        try:
            with open(input_file, 'r', encoding='utf-8') as f:
                self.data = json.load(f)
            
            logger.info(f"Successfully loaded report data from {input_file}")
            logger.debug(f"Data keys: {list(self.data.keys())}")
            
            # Validate data structure
            if not isinstance(self.data, dict):
                raise ValueError(f"Expected dictionary, got {type(self.data)}")
            
            # Log data expectations
            expected_keys = ['summary', 'detailed_analysis', 'date_range', 'formulas']
            missing_keys = [key for key in expected_keys if key not in self.data]
            if missing_keys:
                logger.warning(f"Missing expected data keys: {missing_keys}")
            
            return self.data
            
        except FileNotFoundError:
            logger.error(f"Input file not found: {input_file}")
            raise Exception(f"Failed to load data: Input file not found: {input_file}")
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON in {input_file}: {e}")
            raise Exception(f"Failed to load data: Invalid JSON format: {str(e)}")
        except Exception as e:
            logger.error(f"Unexpected error loading data from {input_file}: {e}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise Exception(f"Failed to load data: {str(e)}")
    
    def load_ml_insights(self, insights_file: str):
        """Load ML insights from JSON file"""
        logger.info(f"Loading ML insights from {insights_file}")
        try:
            with open(insights_file, 'r', encoding='utf-8') as f:
                self.ml_insights = json.load(f)
            
            logger.info(f"Successfully loaded ML insights from {insights_file}")
            logger.debug(f"ML insights keys: {list(self.ml_insights.keys()) if self.ml_insights else 'None'}")
            
            return self.ml_insights
            
        except FileNotFoundError:
            self.ml_insights = None
            logger.warning(f"ML insights file not found: {insights_file}")
        except json.JSONDecodeError as e:
            self.ml_insights = None
            logger.warning(f"Invalid JSON in ML insights file {insights_file}: {e}")
        except Exception as e:
            self.ml_insights = None
            logger.warning(f"Could not load ML insights from {insights_file}: {str(e)}")
    
    def create_summary_sheet(self):
        """Create executive summary sheet"""
        logger.info("Creating executive summary sheet")
        try:
            if 'Sheet' in self.wb.sheetnames:
                self.wb.remove(self.wb['Sheet'])
                logger.debug("Removed default 'Sheet' worksheet")
            
            ws = self.wb.create_sheet('Executive Summary', 0)
            logger.debug("Created 'Executive Summary' worksheet")
            
            # Title
            ws['A1'] = 'Stale PR Audit Report - Executive Summary'
            ws['A1'].font = Font(size=16, bold=True, color='2E3440')
            ws.merge_cells('A1:F1')
            
            # Analysis details
            date_range = self.data.get('date_range', {})
            ws['A3'] = 'Analysis Target:'
            ws['B3'] = str(date_range.get('analysis_target', 'Unknown'))
            ws['A4'] = 'Target Type:'
            ws['B4'] = str(date_range.get('target_type', 'Unknown'))
            ws['A5'] = 'Date Range:'
            ws['B5'] = f"{date_range.get('start_date', 'Unknown')} to {date_range.get('end_date', 'Unknown')}"
            
            logger.debug(f"Added analysis details: target={date_range.get('analysis_target')}, type={date_range.get('target_type')}")
            
            # Key metrics
            row = 7
            ws[f'A{row}'] = 'Key Metrics'
            ws[f'A{row}'].font = self.styles['subheader']
            ws[f'A{row}'].fill = self.styles['subheader_fill']
            ws.merge_cells(f'A{row}:C{row}')
            
            summary = self.data.get('summary', {})
            logger.debug(f"Summary data keys: {list(summary.keys())}")
            
            metrics = [
                ('Total PRs', 'total_prs', safe_int(summary.get('total_prs', 0))),
                ('Open PRs', 'open_prs', safe_int(summary.get('open_prs', 0))),
                ('Inactive PRs', 'inactive_prs', safe_int(summary.get('inactive_prs', 0))),
                ('Active PRs', 'active_prs', safe_int(summary.get('active_prs', 0))),
                ('Abandonment Rate (%)', 'abandonment_rate', safe_percentage(summary.get('abandonment_rate', 0))),
                ('Avg Inactive Days', 'avg_inactive_days', safe_float(summary.get('avg_inactive_days', 0))),
                ('Max Inactive Days', 'max_inactive_days', safe_int(summary.get('max_inactive_days', 0)))
            ]
            
            logger.debug("Processing key metrics:")
            for i, (label, key, value) in enumerate(metrics, 1):
                metric_row = row + i
                ws[f'A{metric_row}'] = label
                ws[f'B{metric_row}'] = value
                ws[f'B{metric_row}'].font = Font(bold=True)
                
                logger.debug(f"  {label}: {value} (from key: {key})")
                
                # Format percentage for abandonment rate
                if key == 'abandonment_rate':
                    try:
                        ws[f'B{metric_row}'].number_format = '0.0%'
                        logger.debug(f"Applied percentage format to {label}")
                    except Exception as e:
                        logger.warning(f"Could not apply percentage format to {label}: {e}")
            
            # Key findings
            findings_row = row + len(metrics) + 2
            ws[f'A{findings_row}'] = 'Key Findings'
            ws[f'A{findings_row}'].font = self.styles['subheader']
            ws[f'A{findings_row}'].fill = self.styles['subheader_fill']
            ws.merge_cells(f'A{findings_row}:F{findings_row}')
            
            findings = summary.get('key_findings', [])
            logger.debug(f"Adding {len(findings)} key findings")
            for i, finding in enumerate(findings, 1):
                finding_row = findings_row + i
                ws[f'A{finding_row}'] = f"• {finding}"
                ws[f'A{finding_row}'].alignment = self.styles['left']
            
            # Apply borders and formatting
            self.apply_sheet_formatting(ws, len(metrics) + len(findings) + 10)
            
            # Auto-adjust column widths
            self.auto_adjust_columns(ws)
            
            logger.info("Executive summary sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating summary sheet: {e}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise Exception(f"Failed to create summary sheet: {str(e)}")
    
    def create_detailed_analysis_sheet(self):
        """Create detailed PR analysis sheet"""
        logger.info("Creating detailed analysis sheet")
        try:
            ws = self.wb.create_sheet('Detailed Analysis')
            
            # Extract PR data
            detailed_analysis = self.data.get('detailed_analysis', {})
            prs = detailed_analysis.get('pull_requests', [])
            
            logger.debug(f"Found {len(prs)} pull requests for detailed analysis")
            
            if not prs:
                ws['A1'] = 'No pull request data available'
                logger.warning("No pull request data available for detailed analysis")
                return
            
            # Create DataFrame
            pr_data = []
            logger.debug("Processing pull request data:")
            
            for i, pr in enumerate(prs):
                try:
                    # Safely extract data with proper type conversion
                    title = str(pr.get('title', ''))
                    title_display = (title[:50] + '...') if len(title) > 50 else title
                    
                    # Extract nested data safely
                    user_login = pr.get('user', {}).get('login', '') if isinstance(pr.get('user'), dict) else ''
                    repo_name = (pr.get('repository_name', '') or 
                               pr.get('base', {}).get('repo', {}).get('full_name', '') if isinstance(pr.get('base'), dict) else '')
                    
                    # Extract dates
                    created_at = pr.get('created_at', '')
                    created_date = created_at.split('T')[0] if created_at and 'T' in created_at else str(created_at)
                    
                    updated_at = pr.get('updated_at', '')
                    updated_date = updated_at.split('T')[0] if updated_at and 'T' in updated_at else str(updated_at)
                    
                    # Extract inactivity data
                    inactivity_duration = pr.get('inactivity_duration', {})
                    inactive_days = safe_int(inactivity_duration.get('days', 0)) if isinstance(inactivity_duration, dict) else 0
                    
                    # Extract analysis data
                    inactivity_analysis = pr.get('inactivity_analysis', {})
                    inactivity_reason = str(inactivity_analysis.get('reason', '')) if isinstance(inactivity_analysis, dict) else ''
                    priority = str(inactivity_analysis.get('priority', '')) if isinstance(inactivity_analysis, dict) else ''
                    
                    # Extract details
                    details = pr.get('details', {})
                    if isinstance(details, dict):
                        review_count = safe_int(details.get('review_count', 0))
                        comment_count = safe_int(details.get('comment_count', 0))
                        failing_checks = safe_int(details.get('failing_checks', 0))
                    else:
                        review_count = comment_count = failing_checks = 0
                    
                    pr_data.append({
                        'PR Number': safe_int(pr.get('number', 0)),
                        'Title': title_display,
                        'Author': str(user_login),
                        'State': str(pr.get('state', '')),
                        'Repository': str(repo_name),
                        'Created Date': created_date,
                        'Updated Date': updated_date,
                        'Inactive Days': inactive_days,
                        'Inactivity Reason': inactivity_reason,
                        'Priority': priority,
                        'Review Count': review_count,
                        'Comment Count': comment_count,
                        'Failing Checks': failing_checks,
                        'Engagement Score': review_count + comment_count,
                        'URL': str(pr.get('html_url', ''))
                    })
                    
                    if i < 5:  # Log first 5 PRs for debugging
                        logger.debug(f"  PR #{safe_int(pr.get('number', 0))}: {title_display[:30]}...")
                
                except Exception as e:
                    logger.warning(f"Error processing PR {i}: {e}")
                    continue
            
            if not pr_data:
                ws['A1'] = 'Failed to process pull request data'
                logger.error("Failed to process any pull request data")
                return
            
            df = pd.DataFrame(pr_data)
            logger.info(f"Created DataFrame with {len(df)} rows and {len(df.columns)} columns")
            
            # Add DataFrame to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Style the header row
            for cell in ws[1]:
                cell.font = self.styles['header']
                cell.fill = self.styles['header_fill']
                cell.alignment = self.styles['center']
                cell.border = self.styles['border']
            
            # Add conditional formatting
            self.add_conditional_formatting(ws, df)
            
            # Add formulas for summary calculations
            last_row = len(df) + 1
            summary_row = last_row + 2
            
            ws[f'A{summary_row}'] = 'Summary Calculations:'
            ws[f'A{summary_row}'].font = self.styles['subheader']
            
            # Add summary formulas with error handling
            formulas = [
                ('Total PRs', f'=COUNTA(A2:A{last_row})'),
                ('Avg Inactive Days', f'=AVERAGE(H2:H{last_row})'),
                ('Max Inactive Days', f'=MAX(H2:H{last_row})'),
                ('High Priority Count', f'=COUNTIF(J2:J{last_row},"high")'),
                ('Avg Engagement Score', f'=AVERAGE(N2:N{last_row})')
            ]
            
            logger.debug("Adding summary formulas:")
            for i, (label, formula) in enumerate(formulas):
                try:
                    row_num = summary_row + i + 1
                    ws[f'A{row_num}'] = label
                    ws[f'B{row_num}'] = formula
                    ws[f'B{row_num}'].font = Font(bold=True)
                    logger.debug(f"  {label}: {formula}")
                except Exception as e:
                    logger.warning(f"Error adding formula {label}: {e}")
            
            # Auto-adjust column widths
            self.auto_adjust_columns(ws)
            
            logger.info("Detailed analysis sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating detailed analysis sheet: {e}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise Exception(f"Failed to create detailed analysis sheet: {str(e)}")
    
    def create_category_breakdown_sheet(self):
        """Create inactivity category breakdown sheet"""
        logger.info("Creating category breakdown sheet")
        try:
            ws = self.wb.create_sheet('Category Breakdown')
            
            detailed_analysis = self.data.get('detailed_analysis', {})
            categories = detailed_analysis.get('inactivity_categories', {})
            
            logger.debug(f"Found {len(categories)} inactivity categories: {list(categories.keys())}")
            
            if not categories:
                ws['A1'] = 'No category data available'
                logger.warning("No category data available")
                return
            
            # Create summary table
            ws['A1'] = 'Inactivity Category Analysis'
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:E1')
            
            # Headers
            headers = ['Category', 'Count', 'Percentage', 'Sample PRs', 'Average Inactive Days']
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=i, value=header)
                cell.font = self.styles['header']
                cell.fill = self.styles['header_fill']
                cell.alignment = self.styles['center']
                cell.border = self.styles['border']
            
            # Calculate total for percentage calculations
            total_prs = sum(safe_int(data.get('count', 0)) for data in categories.values())
            if total_prs == 0:
                total_prs = 1  # Prevent division by zero
            
            logger.debug(f"Total PRs across all categories: {total_prs}")
            
            # Data rows
            row = 4
            for category, data in categories.items():
                try:
                    count = safe_int(data.get('count', 0))
                    percentage = safe_float((count / total_prs) if total_prs > 0 else 0)
                    
                    # Sample PRs
                    sample_prs_data = data.get('prs', [])
                    sample_prs = ', '.join([f"#{safe_int(pr.get('number', 0))}" 
                                          for pr in sample_prs_data[:3] 
                                          if isinstance(pr, dict)])
                    
                    # Calculate average inactive days for this category
                    inactive_days = []
                    for pr in sample_prs_data:
                        if isinstance(pr, dict):
                            inactive_days.append(safe_int(pr.get('inactive_days', 0)))
                    
                    avg_inactive = safe_float(sum(inactive_days) / len(inactive_days) if inactive_days else 0)
                    
                    ws[f'A{row}'] = str(category.replace('_', ' ').title())
                    ws[f'B{row}'] = count
                    ws[f'C{row}'] = percentage
                    try:
                        ws[f'C{row}'].number_format = '0.0%'
                    except Exception as e:
                        logger.warning(f"Could not apply percentage format to category {category}: {e}")
                    ws[f'D{row}'] = sample_prs
                    ws[f'E{row}'] = avg_inactive
                    
                    logger.debug(f"  {category}: {count} PRs ({percentage:.1%})")
                    row += 1
                    
                except Exception as e:
                    logger.warning(f"Error processing category {category}: {e}")
                    continue
            
            # Total row
            ws[f'A{row}'] = 'TOTAL'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = f'=SUM(B4:B{row-1})'
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'] = 1.0  # 100%
            try:
                ws[f'C{row}'].number_format = '0.0%'
            except Exception as e:
                logger.warning(f"Could not apply percentage format to total row: {e}")
            ws[f'C{row}'].font = Font(bold=True)
            
            # Add chart
            if len(categories) > 0:
                self.add_category_chart(ws, len(categories))
            
            # Apply formatting
            for row_num in range(4, row + 1):
                for col_num in range(1, 6):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = self.styles['border']
                    if col_num in [2, 5]:  # Count and average columns
                        cell.alignment = self.styles['center']
            
            # Auto-adjust column widths
            self.auto_adjust_columns(ws)
            
            logger.info("Category breakdown sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating category breakdown sheet: {e}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise Exception(f"Failed to create category breakdown sheet: {str(e)}")
    
    def create_contributor_analysis_sheet(self):
        """Create contributor analysis sheet"""
        logger.info("Creating contributor analysis sheet")
        try:
            ws = self.wb.create_sheet('Contributor Analysis')
            
            detailed_analysis = self.data.get('detailed_analysis', {})
            contributors = detailed_analysis.get('contributor_metrics', {})
            
            logger.debug(f"Found {len(contributors)} contributors: {list(contributors.keys())[:5]}...")
            
            if not contributors:
                ws['A1'] = 'No contributor data available'
                logger.warning("No contributor data available")
                return
            
            # Create summary table
            ws['A1'] = 'Contributor Performance Analysis'
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:F1')
            
            # Headers
            headers = ['Contributor', 'Total PRs', 'Inactive PRs', 'Abandonment Rate (%)', 'Avg Inactive Days', 'Performance Score']
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=i, value=header)
                cell.font = self.styles['header']
                cell.fill = self.styles['header_fill']
                cell.alignment = self.styles['center']
                cell.border = self.styles['border']
            
            # Data rows
            row = 4
            for contributor, stats in contributors.items():
                try:
                    total_prs = safe_int(stats.get('total_prs', 0))
                    inactive_prs = safe_int(stats.get('inactive_prs', 0))
                    avg_inactive_days = safe_float(stats.get('avg_inactive_days', 0))
                    
                    # Calculate abandonment rate
                    abandonment_rate = safe_float((inactive_prs / total_prs * 100) if total_prs > 0 else 0)
                    
                    # Performance score formula (lower abandonment rate and inactive days = higher score)
                    performance_score = safe_float(max(0, 100 - abandonment_rate - (avg_inactive_days / 10)))
                    
                    ws[f'A{row}'] = str(contributor)
                    ws[f'B{row}'] = total_prs
                    ws[f'C{row}'] = inactive_prs
                    ws[f'D{row}'] = abandonment_rate
                    try:
                        ws[f'D{row}'].number_format = '0.0'
                    except Exception as e:
                        logger.warning(f"Could not apply number format to abandonment rate for {contributor}: {e}")
                    ws[f'E{row}'] = avg_inactive_days
                    ws[f'F{row}'] = performance_score
                    try:
                        ws[f'F{row}'].number_format = '0.0'
                    except Exception as e:
                        logger.warning(f"Could not apply number format to performance score for {contributor}: {e}")
                    
                    logger.debug(f"  {contributor}: {total_prs} total, {inactive_prs} inactive, {abandonment_rate:.1f}% abandonment")
                    row += 1
                    
                except Exception as e:
                    logger.warning(f"Error processing contributor {contributor}: {e}")
                    continue
            
            # Add conditional formatting for performance scores
            if row > 4:
                performance_range = f'F4:F{row-1}'
                try:
                    ws.conditional_formatting.add(performance_range, 
                        ColorScaleRule(start_type='min', start_color='BF616A',
                                     mid_type='percentile', mid_value=50, mid_color='EBCB8B',
                                     end_type='max', end_color='A3BE8C'))
                    logger.debug("Applied conditional formatting to performance scores")
                except Exception as e:
                    logger.warning(f"Could not apply conditional formatting: {e}")
            
            # Apply formatting
            for row_num in range(4, row):
                for col_num in range(1, 7):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = self.styles['border']
                    if col_num > 1:  # Numeric columns
                        cell.alignment = self.styles['center']
            
            # Auto-adjust column widths
            self.auto_adjust_columns(ws)
            
            logger.info("Contributor analysis sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating contributor analysis sheet: {e}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise Exception(f"Failed to create contributor analysis sheet: {str(e)}")
    
    def create_ml_insights_sheet(self):
        """Create ML insights sheet if data is available"""
        if not hasattr(self, 'ml_insights') or not self.ml_insights:
            logger.debug("No ML insights data available, skipping ML insights sheet")
            return
        
        logger.info("Creating ML insights sheet")
        try:
            ws = self.wb.create_sheet('ML Insights')
            
            ws['A1'] = 'Machine Learning Analysis Results'
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:F1')
            
            row = 3
            
            # Clustering results
            if 'clustering' in self.ml_insights:
                logger.debug("Adding clustering analysis")
                ws[f'A{row}'] = 'PR Clustering Analysis'
                ws[f'A{row}'].font = self.styles['subheader']
                ws[f'A{row}'].fill = self.styles['subheader_fill']
                ws.merge_cells(f'A{row}:F{row}')
                row += 2
                
                # Cluster headers
                cluster_headers = ['Cluster', 'Size', 'Avg Inactive Days', 'Abandonment Rate', 'Dominant Reason', 'Characteristics']
                for i, header in enumerate(cluster_headers, 1):
                    cell = ws.cell(row=row, column=i, value=header)
                    cell.font = self.styles['header']
                    cell.fill = self.styles['header_fill']
                    cell.alignment = self.styles['center']
                    cell.border = self.styles['border']
                
                row += 1
                
                # Cluster data
                clustering_data = self.ml_insights['clustering']
                for cluster_name, cluster_info in clustering_data.items():
                    try:
                        ws[f'A{row}'] = str(cluster_name.replace('cluster_', 'Cluster ').title())
                        ws[f'B{row}'] = safe_int(cluster_info.get('size', 0))
                        ws[f'C{row}'] = round(safe_float(cluster_info.get('avg_inactive_days', 0)), 1)
                        
                        abandonment_rate = safe_float(cluster_info.get('abandonment_rate', 0))
                        ws[f'D{row}'] = f"{round(abandonment_rate * 100, 1)}%"
                        
                        ws[f'E{row}'] = str(cluster_info.get('dominant_reason', 'Unknown')).title()
                        
                        characteristics = cluster_info.get('characteristics', [])
                        ws[f'F{row}'] = ', '.join([str(c) for c in characteristics])
                        
                        # Apply borders
                        for col in range(1, 7):
                            ws.cell(row=row, column=col).border = self.styles['border']
                        
                        row += 1
                        
                    except Exception as e:
                        logger.warning(f"Error processing cluster {cluster_name}: {e}")
                        continue
                
                row += 2
            
            # Risk analysis
            if 'risk_analysis' in self.ml_insights:
                logger.debug("Adding risk analysis")
                ws[f'A{row}'] = 'Risk Analysis Summary'
                ws[f'A{row}'].font = self.styles['subheader']
                ws[f'A{row}'].fill = self.styles['subheader_fill']
                ws.merge_cells(f'A{row}:D{row}')
                row += 2
                
                risk_data = self.ml_insights['risk_analysis']
                
                # Risk distribution
                if 'risk_distribution' in risk_data:
                    ws[f'A{row}'] = 'Risk Level Distribution:'
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1
                    
                    for risk_level, count in risk_data['risk_distribution'].items():
                        ws[f'A{row}'] = f"{risk_level} Risk:"
                        ws[f'B{row}'] = safe_int(count)
                        row += 1
                    
                    row += 1
                
                # High-risk PRs
                if 'high_risk_prs' in risk_data and risk_data['high_risk_prs']:
                    ws[f'A{row}'] = 'High-Risk PRs (Top 10):'
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1
                    
                    # Headers for high-risk PRs
                    risk_headers = ['PR Number', 'Risk Score', 'Category']
                    for i, header in enumerate(risk_headers, 1):
                        cell = ws.cell(row=row, column=i, value=header)
                        cell.font = self.styles['header']
                        cell.fill = self.styles['header_fill']
                        cell.border = self.styles['border']
                    
                    row += 1
                    
                    # High-risk PR data
                    for pr in risk_data['high_risk_prs'][:10]:
                        try:
                            ws[f'A{row}'] = f"#{safe_int(pr.get('pr_number', 0))}"
                            ws[f'B{row}'] = round(safe_float(pr.get('risk_score', 0)), 1)
                            ws[f'C{row}'] = str(pr.get('reason_category', 'Unknown'))
                            
                            # Apply borders
                            for col in range(1, 4):
                                ws.cell(row=row, column=col).border = self.styles['border']
                            
                            row += 1
                            
                        except Exception as e:
                            logger.warning(f"Error processing high-risk PR: {e}")
                            continue
            
            # Recommendations
            if 'recommendations' in self.ml_insights:
                logger.debug("Adding ML recommendations")
                row += 2
                ws[f'A{row}'] = 'ML-Generated Recommendations'
                ws[f'A{row}'].font = self.styles['subheader']
                ws[f'A{row}'].fill = self.styles['subheader_fill']
                ws.merge_cells(f'A{row}:F{row}')
                row += 2
                
                # Recommendation headers
                rec_headers = ['Priority', 'Category', 'Recommendation']
                for i, header in enumerate(rec_headers, 1):
                    cell = ws.cell(row=row, column=i, value=header)
                    cell.font = self.styles['header']
                    cell.fill = self.styles['header_fill']
                    cell.border = self.styles['border']
                
                row += 1
                
                # Recommendation data
                recommendations = self.ml_insights['recommendations']
                for rec in recommendations[:10]:
                    try:
                        ws[f'A{row}'] = str(rec.get('priority', 'Medium'))
                        ws[f'B{row}'] = str(rec.get('category', 'General'))
                        ws[f'C{row}'] = str(rec.get('suggestion', ''))
                        
                        # Apply borders and wrap text
                        for col in range(1, 4):
                            cell = ws.cell(row=row, column=col)
                            cell.border = self.styles['border']
                            if col == 3:  # Recommendation text
                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                        
                        row += 1
                        
                    except Exception as e:
                        logger.warning(f"Error processing recommendation: {e}")
                        continue
            
            # Auto-adjust column widths
            self.auto_adjust_columns(ws)
            
            logger.info("ML insights sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating ML insights sheet: {e}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise Exception(f"Failed to create ML insights sheet: {str(e)}")
    
    def create_formulas_sheet(self):
        """Create formulas documentation sheet"""
        logger.info("Creating formulas documentation sheet")
        try:
            ws = self.wb.create_sheet('Formulas')
            
            ws['A1'] = 'Formula Documentation'
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:D1')
            
            # Headers
            headers = ['Formula Name', 'Formula', 'Description', 'Variables']
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=i, value=header)
                cell.font = self.styles['header']
                cell.fill = self.styles['header_fill']
                cell.alignment = self.styles['center']
                cell.border = self.styles['border']
            
            # Formula data
            formulas = self.data.get('formulas', {})
            logger.debug(f"Found {len(formulas)} formulas: {list(formulas.keys())}")
            
            formula_descriptions = {
                'abandonment_rate': 'Percentage of PRs that are inactive',
                'inactivity_duration': 'Time since last activity on PR',
                'avg_inactive_days': 'Average days PRs remain inactive',
                'max_inactive_days': 'Maximum inactive period observed',
                'failing_ci_rate': 'Percentage of PRs with failing CI',
                'no_review_rate': 'Percentage of PRs without reviews',
                'outdated_rate': 'Percentage of PRs that are outdated',
                'abandoned_rate': 'Percentage of completely abandoned PRs',
                'active_rate': 'Percentage of actively maintained PRs'
            }
            
            formula_variables = {
                'abandonment_rate': 'INACTIVE_PRS, TOTAL_PRS',
                'inactivity_duration': 'CURRENT_TIME, LAST_ACTIVITY_TIME',
                'avg_inactive_days': 'INACTIVE_DAYS array, COUNT',
                'max_inactive_days': 'INACTIVE_DAYS array',
                'failing_ci_rate': 'FAILING_CI_PRS, TOTAL_PRS',
                'no_review_rate': 'NO_REVIEW_PRS, TOTAL_PRS',
                'outdated_rate': 'OUTDATED_PRS, TOTAL_PRS',
                'abandoned_rate': 'ABANDONED_PRS, TOTAL_PRS',
                'active_rate': 'ACTIVE_PRS, TOTAL_PRS'
            }
            
            row = 4
            for formula_name, formula_text in formulas.items():
                try:
                    ws[f'A{row}'] = str(formula_name.replace('_', ' ').title())
                    ws[f'B{row}'] = str(formula_text)
                    ws[f'C{row}'] = formula_descriptions.get(formula_name, 'Custom calculation')
                    ws[f'D{row}'] = formula_variables.get(formula_name, 'Various')
                    
                    # Apply borders and formatting
                    for col in range(1, 5):
                        cell = ws.cell(row=row, column=col)
                        cell.border = self.styles['border']
                        if col == 2:  # Formula column
                            cell.font = Font(family='Courier New')
                        if col in [3, 4]:  # Description columns
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    logger.debug(f"  Added formula: {formula_name}")
                    row += 1
                    
                except Exception as e:
                    logger.warning(f"Error processing formula {formula_name}: {e}")
                    continue
            
            # Auto-adjust column widths
            column_widths = [20, 30, 40, 25]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            logger.info("Formulas documentation sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating formulas sheet: {e}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise Exception(f"Failed to create formulas sheet: {str(e)}")
    
    def add_conditional_formatting(self, ws, df):
        """Add conditional formatting to the worksheet"""
        if len(df) == 0:
            logger.debug("No data for conditional formatting")
            return
        
        last_row = len(df) + 1
        logger.debug(f"Applying conditional formatting to {last_row - 1} rows")
        
        try:
            # Inactive days - color scale (red for high values)
            inactive_range = f'H2:H{last_row}'
            ws.conditional_formatting.add(inactive_range, 
                ColorScaleRule(start_type='num', start_value=0, start_color='A3BE8C',
                             mid_type='num', mid_value=30, mid_color='EBCB8B',
                             end_type='max', end_color='BF616A'))
            logger.debug("Applied color scale to inactive days column")
            
            # Priority - color coding
            priority_range = f'J2:J{last_row}'
            priority_rules = [
                ('critical', 'BF616A'),
                ('high', 'D08770'),
                ('medium', 'EBCB8B'),
                ('low', 'A3BE8C')
            ]
            
            for priority, color in priority_rules:
                try:
                    ws.conditional_formatting.add(priority_range,
                        CellIsRule(operator='equal', formula=[f'"{priority}"'], 
                                  fill=PatternFill(start_color=color, end_color=color, fill_type='solid')))
                except Exception as e:
                    logger.warning(f"Could not apply conditional formatting for priority {priority}: {e}")
            
            logger.debug("Applied priority color coding")
            
        except Exception as e:
            logger.warning(f"Could not apply conditional formatting: {e}")
    
    def add_category_chart(self, ws, num_categories):
        """Add a pie chart for category breakdown"""
        try:
            chart = PieChart()
            chart.title = "Inactivity Categories Distribution"
            
            # Data for the chart
            data = Reference(ws, min_col=2, min_row=4, max_row=4 + num_categories - 1)
            categories = Reference(ws, min_col=1, min_row=4, max_row=4 + num_categories - 1)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(categories)
            
            # Position the chart
            ws.add_chart(chart, "H4")
            logger.debug("Added pie chart for category distribution")
            
        except Exception as e:
            logger.warning(f"Could not create category chart: {e}")
    
    def apply_sheet_formatting(self, ws, max_row):
        """Apply general formatting to a sheet"""
        try:
            # Apply borders to used range
            for row in range(1, max_row + 1):
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell.border = self.styles['border']
            logger.debug(f"Applied formatting to sheet with {max_row} rows")
        except Exception as e:
            logger.warning(f"Could not apply sheet formatting: {e}")
    
    def auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            logger.debug("Auto-adjusted column widths")
        except Exception as e:
            logger.warning(f"Could not auto-adjust columns: {e}")
    
    def generate_excel_report(self, output_file: str) -> str:
        """Generate the complete Excel report"""
        logger.info(f"Starting Excel report generation to {output_file}")
        
        try:
            # Validate data before processing
            if not self.data:
                raise ValueError("No data loaded for report generation")
            
            logger.info("Data validation passed, creating sheets...")
            
            # Create all sheets
            logger.info("Creating summary sheet...")
            self.create_summary_sheet()
            
            logger.info("Creating detailed analysis sheet...")
            self.create_detailed_analysis_sheet()
            
            logger.info("Creating category breakdown sheet...")
            self.create_category_breakdown_sheet()
            
            logger.info("Creating contributor analysis sheet...")
            self.create_contributor_analysis_sheet()
            
            if hasattr(self, 'ml_insights') and self.ml_insights:
                logger.info("Creating ML insights sheet...")
                self.create_ml_insights_sheet()
            else:
                logger.debug("Skipping ML insights sheet (no data)")
            
            logger.info("Creating formulas documentation sheet...")
            self.create_formulas_sheet()
            
            # Save the workbook
            logger.info(f"Saving workbook to {output_file}")
            self.wb.save(output_file)
            
            logger.info(f"Excel report generated successfully: {output_file}")
            return output_file
            
        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise Exception(f"Failed to generate Excel report: {str(e)}")

def main():
    parser = argparse.ArgumentParser(description='Generate Excel report from Stale PR Audit data')
    parser.add_argument('--input', '-i', required=True, help='Input JSON report file')
    parser.add_argument('--output', '-o', required=True, help='Output Excel file path')
    parser.add_argument('--insights', help='ML insights JSON file (optional)')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose output')
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    
    args = parser.parse_args()
    
    # Set logging level based on arguments
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
        logger.debug("Debug logging enabled")
    elif args.verbose:
        logging.getLogger().setLevel(logging.INFO)
    
    try:
        logger.info("Starting Excel report generation...")
        logger.info(f"Input file: {args.input}")
        logger.info(f"Output file: {args.output}")
        if args.insights:
            logger.info(f"Insights file: {args.insights}")
        
        # Initialize generator
        logger.debug("Initializing Excel generator")
        generator = StalePRExcelGenerator(verbose=args.verbose)
        
        # Load data
        logger.info("Loading report data...")
        generator.load_data(args.input)
        
        # Load ML insights if provided
        if args.insights:
            logger.info("Loading ML insights...")
            generator.load_ml_insights(args.insights)
        else:
            logger.debug("No ML insights file provided")
        
        # Generate report
        logger.info("Generating Excel report...")
        output_file = generator.generate_excel_report(args.output)
        
        print(f"\n✅ Excel report generated successfully: {output_file}")
        print("📋 Report includes:")
        print("  • Executive Summary with key metrics")
        print("  • Detailed PR analysis with conditional formatting")
        print("  • Category breakdown with charts")
        print("  • Contributor performance analysis")
        if args.insights:
            print("  • ML insights and recommendations")
        print("  • Formula documentation")
        
        logger.info("Excel report generation completed successfully")
        
    except Exception as e:
        logger.error(f"Excel report generation failed: {e}")
        print(f"\n❌ Error: Failed to generate Excel report: {str(e)}")
        if args.debug:
            print(f"Traceback: {traceback.format_exc()}")
        sys.exit(1)

if __name__ == "__main__":
    main()