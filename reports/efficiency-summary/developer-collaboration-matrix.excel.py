#!/usr/bin/env python3
"""
Developer Collaboration Matrix - Excel Report Generator

Creates interactive Excel workbooks with multiple sheets, formulas, and visualizations
from JSON/CSV collaboration analysis data.
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Union

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.defined_name import DefinedName

class CollaborationExcelGenerator:
    """Excel report generator for collaboration analysis data."""
    
    def __init__(self, verbose: bool = False):
        self.verbose = verbose
        self.workbook = Workbook()
        self.data = None
        self.insights_data = None
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # Define styles
        self.styles = self._create_styles()
        
    def _create_styles(self) -> Dict[str, Any]:
        """Create consistent styling for the Excel workbook."""
        return {
            'header': Font(name='Segoe UI', size=14, bold=True, color='FFFFFF'),
            'subheader': Font(name='Segoe UI', size=12, bold=True, color='2F4F4F'),
            'body': Font(name='Segoe UI', size=10, color='2F4F4F'),
            'title': Font(name='Segoe UI', size=16, bold=True, color='1F4E79'),
            'header_fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'accent_fill': PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid'),
            'border': Border(
                left=Side(style='thin', color='D0D0D0'),
                right=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='D0D0D0')
            ),
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center'),
            'right': Alignment(horizontal='right', vertical='center')
        }
    
    def load_data(self, file_path: str, insights_path: Optional[str] = None) -> None:
        """Load collaboration data from JSON or CSV files."""
        try:
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.json':
                with open(file_path, 'r', encoding='utf-8') as f:
                    self.data = json.load(f)
            elif file_ext == '.csv':
                # For CSV, create a basic structure
                df = pd.read_csv(file_path)
                self.data = {
                    'summary': {},
                    'detailed_analysis': {'csv_data': df.to_dict('records')},
                    'date_range': {'start_date': 'N/A', 'end_date': 'N/A'}
                }
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            # Load insights if provided
            if insights_path and os.path.exists(insights_path):
                with open(insights_path, 'r', encoding='utf-8') as f:
                    self.insights_data = json.load(f)
            
            if self.verbose:
                print(f"✅ Loaded data from {file_path}")
                if self.insights_data:
                    print(f"✅ Loaded insights from {insights_path}")
                    
        except Exception as e:
            raise Exception(f"Error loading data: {e}")
    
    def create_summary_sheet(self) -> None:
        """Create executive summary sheet."""
        ws = self.workbook.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = 'Developer Collaboration Matrix - Executive Summary'
        ws['A1'].font = self.styles['title']
        ws.merge_cells('A1:E1')
        
        # Date range
        date_range = self.data.get('date_range', {})
        ws['A3'] = 'Analysis Period:'
        ws['B3'] = f"{date_range.get('start_date', 'N/A')} to {date_range.get('end_date', 'N/A')}"
        ws['A3'].font = self.styles['subheader']
        
        # Target info
        ws['A4'] = 'Analysis Target:'
        ws['B4'] = f"{date_range.get('analysis_target', 'N/A')} ({date_range.get('analysis_mode', 'N/A')})"
        
        # Key metrics
        summary = self.data.get('summary', {})
        total = self.data.get('total', {})
        
        metrics = [
            ('Total Pull Requests', summary.get('total_pull_requests', 0)),
            ('Total Collaborators', summary.get('total_collaborators', 0)),
            ('Total Interactions', summary.get('total_interactions', 0)),
            ('Discussion Threads', summary.get('total_discussion_threads', 0)),
            ('Average Collaboration Score', summary.get('average_collaboration_score', 0)),
            ('Reviews Given', total.get('reviews_given', 0)),
            ('Comments Made', total.get('comments_made', 0)),
            ('PRs Created', total.get('prs_created', 0))
        ]
        
        # Metrics table
        ws['A6'] = 'Key Metrics'
        ws['A6'].font = self.styles['subheader']
        
        headers = ['Metric', 'Value']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.alignment = self.styles['center']
            cell.border = self.styles['border']
        
        for row, (metric, value) in enumerate(metrics, 8):
            ws.cell(row=row, column=1, value=metric).font = self.styles['body']
            value_cell = ws.cell(row=row, column=2, value=float(value) if isinstance(value, (int, float)) else value)
            value_cell.font = self.styles['body']
            value_cell.alignment = self.styles['right']
            
            # Apply borders
            for col in range(1, 3):
                ws.cell(row=row, column=col).border = self.styles['border']
        
        # Highlights section
        ws['D6'] = 'Key Highlights'
        ws['D6'].font = self.styles['subheader']
        
        highlights = [
            f"Most Active: {summary.get('most_active_collaborator', 'N/A')}",
            f"Most Diverse: {summary.get('most_diverse_collaborator', 'N/A')}",
            f"Bottlenecks: {summary.get('collaboration_bottlenecks', 0)}",
            f"Processing Time: {round((summary.get('processing_time_ms', 0)) / 1000, 1)}s"
        ]
        
        for row, highlight in enumerate(highlights, 7):
            ws.cell(row=row, column=4, value=highlight).font = self.styles['body']
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 20
    
    def create_collaboration_matrix_sheet(self) -> None:
        """Create collaboration matrix analysis sheet."""
        ws = self.workbook.create_sheet("Collaboration Matrix")
        
        # Title
        ws['A1'] = 'Collaboration Matrix Analysis'
        ws['A1'].font = self.styles['title']
        ws.merge_cells('A1:F1')
        
        collaboration_data = self.data.get('detailed_analysis', {}).get('collaboration_matrix', {})
        user_stats = collaboration_data.get('user_stats', {})
        interactions = collaboration_data.get('interactions', [])
        
        if not user_stats:
            ws['A3'] = 'No collaboration matrix data available'
            return
        
        # User statistics table
        ws['A3'] = 'User Statistics'
        ws['A3'].font = self.styles['subheader']
        
        headers = ['User', 'PRs Created', 'Reviews Given', 'Comments Made', 'Collaborators']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.alignment = self.styles['center']
            cell.border = self.styles['border']
        
        # User data
        user_data = []
        for user, stats in user_stats.items():
            user_data.append([
                user,
                int(stats.get('prs_created', 0)),
                int(stats.get('reviews_given', 0)),
                int(stats.get('comments_made', 0)),
                int(stats.get('collaborators', 0))
            ])
        
        # Sort by collaboration score if available
        collaboration_scores = self.data.get('detailed_analysis', {}).get('collaboration_scores', {})
        if collaboration_scores:
            user_data.sort(key=lambda x: collaboration_scores.get(x[0], {}).get('collaboration_score', 0), reverse=True)
        
        for row, user_row in enumerate(user_data, 5):
            for col, value in enumerate(user_row, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles['body']
                cell.border = self.styles['border']
                if col > 1:  # Numeric columns
                    cell.alignment = self.styles['right']
        
        # Add conditional formatting for numeric columns
        if len(user_data) > 0:
            for col in range(2, 6):  # PRs, Reviews, Comments, Collaborators
                col_letter = chr(ord('A') + col - 1)
                range_ref = f"{col_letter}5:{col_letter}{4 + len(user_data)}"
                ws.conditional_formatting.add(range_ref, ColorScaleRule(
                    start_type='min', start_color='FFFFFF',
                    mid_type='percentile', mid_value=50, mid_color='92D050',
                    end_type='max', end_color='00B050'
                ))
        
        # Interaction summary
        start_row = 6 + len(user_data)
        ws[f'A{start_row}'] = 'Interaction Summary'
        ws[f'A{start_row}'].font = self.styles['subheader']
        
        interaction_types = {}
        for interaction in interactions:
            int_type = interaction.get('type', 'unknown')
            interaction_types[int_type] = interaction_types.get(int_type, 0) + 1
        
        for i, (int_type, count) in enumerate(interaction_types.items(), start_row + 1):
            ws[f'A{i}'] = int_type.title()
            ws[f'B{i}'] = count
            ws[f'A{i}'].font = self.styles['body']
            ws[f'B{i}'].font = self.styles['body']
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
    
    def create_collaboration_scores_sheet(self) -> None:
        """Create collaboration scores analysis sheet."""
        ws = self.workbook.create_sheet("Collaboration Scores")
        
        # Title
        ws['A1'] = 'Collaboration Scores Analysis'
        ws['A1'].font = self.styles['title']
        ws.merge_cells('A1:G1')
        
        collaboration_scores = self.data.get('detailed_analysis', {}).get('collaboration_scores', {})
        
        if not collaboration_scores:
            ws['A3'] = 'No collaboration scores data available'
            return
        
        # Scores table
        ws['A3'] = 'Collaboration Scores by User'
        ws['A3'].font = self.styles['subheader']
        
        headers = ['User', 'Collaboration Score', 'Diversity Score', 'Activity Score', 'Intensity Score']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.alignment = self.styles['center']
            cell.border = self.styles['border']
        
        # Sort users by collaboration score
        sorted_users = sorted(
            collaboration_scores.items(),
            key=lambda x: x[1].get('collaboration_score', 0),
            reverse=True
        )
        
        for row, (user, scores) in enumerate(sorted_users, 5):
            data = [
                user,
                round(float(scores.get('collaboration_score', 0)), 2),
                round(float(scores.get('diversity_score', 0)), 2),
                round(float(scores.get('activity_score', 0)), 2),
                round(float(scores.get('intensity_score', 0)), 2)
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles['body']
                cell.border = self.styles['border']
                if col > 1:  # Numeric columns
                    cell.alignment = self.styles['right']
        
        # Add conditional formatting
        if len(sorted_users) > 0:
            # Collaboration score column (B)
            range_ref = f"B5:B{4 + len(sorted_users)}"
            ws.conditional_formatting.add(range_ref, ColorScaleRule(
                start_type='min', start_color='FF6B6B',
                mid_type='percentile', mid_value=50, mid_color='FFE66D',
                end_type='max', end_color='4ECDC4'
            ))
            
            # Add data bars for activity score
            range_ref = f"D5:D{4 + len(sorted_users)}"
            ws.conditional_formatting.add(range_ref, DataBarRule(
                start_type='min', start_value=0,
                end_type='max', end_value=None,
                color='4472C4'
            ))
        
        # Statistics section
        start_row = 6 + len(sorted_users)
        ws[f'A{start_row}'] = 'Score Statistics'
        ws[f'A{start_row}'].font = self.styles['subheader']
        
        # Calculate statistics
        scores = [float(s.get('collaboration_score', 0)) for s in collaboration_scores.values()]
        if scores:
            stats = [
                ('Average Score', round(sum(scores) / len(scores), 2)),
                ('Highest Score', round(max(scores), 2)),
                ('Lowest Score', round(min(scores), 2)),
                ('Score Range', round(max(scores) - min(scores), 2))
            ]
            
            for i, (stat_name, value) in enumerate(stats, start_row + 1):
                ws[f'A{i}'] = stat_name
                ws[f'B{i}'] = value
                ws[f'A{i}'].font = self.styles['body']
                ws[f'B{i}'].font = self.styles['body']
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18
    
    def create_temporal_analysis_sheet(self) -> None:
        """Create temporal patterns analysis sheet."""
        ws = self.workbook.create_sheet("Temporal Analysis")
        
        # Title
        ws['A1'] = 'Temporal Collaboration Patterns'
        ws['A1'].font = self.styles['title']
        ws.merge_cells('A1:F1')
        
        temporal_data = self.data.get('detailed_analysis', {}).get('temporal_analysis', {})
        
        if not temporal_data:
            ws['A3'] = 'No temporal analysis data available'
            return
        
        # Monthly trends
        monthly_data = temporal_data.get('by_month', {})
        if monthly_data:
            ws['A3'] = 'Monthly Activity'
            ws['A3'].font = self.styles['subheader']
            
            # Headers
            ws['A4'] = 'Month'
            ws['B4'] = 'PR Count'
            ws['A4'].font = self.styles['header']
            ws['B4'].font = self.styles['header']
            
            sorted_months = sorted(monthly_data.keys())
            for row, month in enumerate(sorted_months, 5):
                ws[f'A{row}'] = month
                ws[f'B{row}'] = monthly_data[month]
                ws[f'A{row}'].font = self.styles['body']
                ws[f'B{row}'].font = self.styles['body']
        
        # Weekly patterns
        weekly_data = temporal_data.get('by_day_of_week', {})
        if weekly_data:
            start_col = 'D'
            ws[f'{start_col}3'] = 'Weekly Patterns'
            ws[f'{start_col}3'].font = self.styles['subheader']
            
            # Headers
            ws[f'{start_col}4'] = 'Day'
            ws[f'{chr(ord(start_col) + 1)}4'] = 'Activity'
            ws[f'{start_col}4'].font = self.styles['header']
            ws[f'{chr(ord(start_col) + 1)}4'].font = self.styles['header']
            
            day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            for row, day in enumerate(day_order, 5):
                if day in weekly_data:
                    ws[f'{start_col}{row}'] = day
                    ws[f'{chr(ord(start_col) + 1)}{row}'] = weekly_data[day]
                    ws[f'{start_col}{row}'].font = self.styles['body']
                    ws[f'{chr(ord(start_col) + 1)}{row}'].font = self.styles['body']
        
        # Hourly distribution
        hourly_data = temporal_data.get('by_hour', {})
        if hourly_data:
            start_row = 15
            ws[f'A{start_row}'] = 'Hourly Distribution'
            ws[f'A{start_row}'].font = self.styles['subheader']
            
            # Convert to sorted list
            hourly_list = [(int(hour), count) for hour, count in hourly_data.items() if hour.isdigit()]
            hourly_list.sort()
            
            # Create hourly table (in rows of 6)
            for i, (hour, count) in enumerate(hourly_list):
                row = start_row + 2 + (i // 6)
                col = 1 + ((i % 6) * 2)
                
                ws.cell(row=row, column=col, value=f"{hour:02d}:00").font = self.styles['body']
                ws.cell(row=row, column=col + 1, value=count).font = self.styles['body']
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 12
    
    def create_ml_insights_sheet(self) -> None:
        """Create ML insights sheet if data is available."""
        if not self.insights_data:
            return
        
        ws = self.workbook.create_sheet("ML Insights")
        
        # Title
        ws['A1'] = 'Machine Learning Insights'
        ws['A1'].font = self.styles['title']
        ws.merge_cells('A1:F1')
        
        # Metadata
        metadata = self.insights_data.get('metadata', {})
        ws['A3'] = 'Analysis Information'
        ws['A3'].font = self.styles['subheader']
        
        info_items = [
            ('Analysis Date', metadata.get('analysis_timestamp', 'N/A')[:19]),
            ('Users Analyzed', metadata.get('total_users_analyzed', 0)),
            ('ML Algorithm', metadata.get('ml_model_info', {}).get('clustering_algorithm', 'N/A')),
            ('Features Used', metadata.get('ml_model_info', {}).get('feature_count', 0))
        ]
        
        for row, (label, value) in enumerate(info_items, 4):
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.styles['body']
            ws[f'B{row}'].font = self.styles['body']
        
        # Clustering results
        clustering_data = self.insights_data.get('clustering_analysis', {})
        clusters = clustering_data.get('clusters', [])
        
        if clusters:
            start_row = 9
            ws[f'A{start_row}'] = 'Cluster Analysis'
            ws[f'A{start_row}'].font = self.styles['subheader']
            
            # Cluster headers
            headers = ['Cluster ID', 'Size', 'Collaboration Level', 'Avg Score']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row + 1, column=col, value=header)
                cell.font = self.styles['header']
                cell.fill = self.styles['header_fill']
                cell.border = self.styles['border']
            
            # Cluster data
            for row, cluster in enumerate(clusters, start_row + 2):
                characteristics = cluster.get('characteristics', {})
                data = [
                    cluster.get('cluster_id', 0),
                    cluster.get('size', 0),
                    characteristics.get('collaboration_level', 'Unknown'),
                    round(float(cluster.get('avg_collaboration_score', 0)), 2)
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = self.styles['body']
                    cell.border = self.styles['border']
        
        # Recommendations
        recommendations = self.insights_data.get('collaboration_recommendations', [])
        if recommendations:
            start_row = 15
            ws[f'A{start_row}'] = 'ML Recommendations'
            ws[f'A{start_row}'].font = self.styles['subheader']
            
            for row, rec in enumerate(recommendations[:10], start_row + 1):  # Limit to 10
                priority_icon = '🔴' if rec.get('priority') == 'high' else '🟡' if rec.get('priority') == 'medium' else '🟢'
                ws[f'A{row}'] = f"{priority_icon} {rec.get('type', '').replace('_', ' ').title()}"
                ws[f'B{row}'] = rec.get('description', '')
                ws[f'A{row}'].font = self.styles['body']
                ws[f'B{row}'].font = self.styles['body']
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def create_formulas_sheet(self) -> None:
        """Create formulas reference sheet."""
        ws = self.workbook.create_sheet("Formulas")
        
        # Title
        ws['A1'] = 'Calculation Formulas Reference'
        ws['A1'].font = self.styles['title']
        ws.merge_cells('A1:D1')
        
        formulas = self.data.get('formulas', {})
        
        if not formulas:
            ws['A3'] = 'No formulas data available'
            return
        
        # Headers
        ws['A3'] = 'Formula Name'
        ws['B3'] = 'Formula'
        ws['C3'] = 'Description'
        ws['A3'].font = self.styles['header']
        ws['B3'].font = self.styles['header']
        ws['C3'].font = self.styles['header']
        ws['A3'].fill = self.styles['header_fill']
        ws['B3'].fill = self.styles['header_fill']
        ws['C3'].fill = self.styles['header_fill']
        
        # Formula descriptions (mapping formulas to descriptions)
        formula_descriptions = {
            'collaboration_score': 'Overall collaboration effectiveness combining diversity, activity, and intensity',
            'diversity_score': 'Number of unique collaborators a user interacts with',
            'activity_score': 'Total contribution activity including PRs, reviews, and comments',
            'intensity_score': 'Weighted sum of interaction values based on interaction type',
            'review_weight': 'Scoring system for different types of code reviews',
            'interaction_frequency': 'Count of interactions between specific user pairs',
            'discussion_thread_size': 'Number of participants in each discussion thread',
            'bottleneck_threshold': 'Criteria for identifying collaboration bottlenecks',
            'average_collaboration_score': 'Mean collaboration score across all team members',
            'temporal_activity': 'Activity distribution across time periods'
        }
        
        row = 4
        for formula_name, formula_text in formulas.items():
            ws[f'A{row}'] = formula_name.replace('_', ' ').title()
            ws[f'B{row}'] = str(formula_text)
            ws[f'C{row}'] = formula_descriptions.get(formula_name, 'Calculation formula used in analysis')
            
            ws[f'A{row}'].font = self.styles['body']
            ws[f'B{row}'].font = Font(name='Consolas', size=9, color='2F4F4F')  # Monospace for formulas
            ws[f'C{row}'].font = self.styles['body']
            
            # Add borders
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = self.styles['border']
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 60
    
    def create_charts(self) -> None:
        """Add charts to relevant sheets."""
        try:
            # Chart 1: Collaboration Scores (if data exists)
            if 'Collaboration Scores' in self.workbook.sheetnames:
                ws = self.workbook['Collaboration Scores']
                
                # Find data range
                collaboration_scores = self.data.get('detailed_analysis', {}).get('collaboration_scores', {})
                if collaboration_scores:
                    data_rows = len(collaboration_scores)
                    
                    # Create bar chart for top 10 collaboration scores
                    chart = BarChart()
                    chart.title = "Top Collaboration Scores"
                    chart.y_axis.title = 'Collaboration Score'
                    chart.x_axis.title = 'Users'
                    
                    # Data references
                    data_ref = Reference(ws, min_col=2, min_row=4, max_col=2, max_row=min(14, 4 + data_rows))  # Top 10
                    labels_ref = Reference(ws, min_col=1, min_row=5, max_row=min(14, 4 + data_rows))
                    
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(labels_ref)
                    
                    # Position chart
                    ws.add_chart(chart, "G5")
            
            # Chart 2: Temporal Analysis
            if 'Temporal Analysis' in self.workbook.sheetnames:
                ws = self.workbook['Temporal Analysis']
                
                temporal_data = self.data.get('detailed_analysis', {}).get('temporal_analysis', {})
                monthly_data = temporal_data.get('by_month', {})
                
                if monthly_data and len(monthly_data) > 1:
                    # Create line chart for monthly trends
                    chart = LineChart()
                    chart.title = "Monthly Activity Trends"
                    chart.y_axis.title = 'PR Count'
                    chart.x_axis.title = 'Month'
                    
                    # Find data range
                    data_rows = len(monthly_data)
                    data_ref = Reference(ws, min_col=2, min_row=4, max_col=2, max_row=4 + data_rows)
                    labels_ref = Reference(ws, min_col=1, min_row=5, max_row=4 + data_rows)
                    
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(labels_ref)
                    
                    # Position chart
                    ws.add_chart(chart, "G5")
            
        except Exception as e:
            if self.verbose:
                print(f"⚠️ Warning: Could not create charts: {e}")
    
    def add_named_ranges(self) -> None:
        """Add named ranges for key data areas."""
        try:
            # Define named ranges for key metrics that can be referenced in formulas
            
            # Summary metrics
            if 'Summary' in self.workbook.sheetnames:
                ws = self.workbook['Summary']
                
                # Create named range for key metrics
                defined_name = DefinedName('TotalPRs', attr_text=f"Summary!$B$8")
                self.workbook.defined_names.append(defined_name)
                
                defined_name = DefinedName('TotalCollaborators', attr_text=f"Summary!$B$9")
                self.workbook.defined_names.append(defined_name)
                
                defined_name = DefinedName('AvgCollaborationScore', attr_text=f"Summary!$B$12")
                self.workbook.defined_names.append(defined_name)
            
            if self.verbose:
                print("✅ Added named ranges for key metrics")
                
        except Exception as e:
            if self.verbose:
                print(f"⚠️ Warning: Could not create named ranges: {e}")
    
    def save_workbook(self, output_path: str) -> str:
        """Save the Excel workbook."""
        try:
            # Ensure directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
            
            # Save workbook
            self.workbook.save(output_path)
            
            if self.verbose:
                print(f"✅ Excel report saved: {output_path}")
            
            return output_path
            
        except Exception as e:
            raise Exception(f"Failed to save Excel file: {e}")
    
    def generate_report(self, output_path: str) -> str:
        """Generate complete Excel report."""
        if not self.data:
            raise ValueError("No data loaded. Call load_data() first.")
        
        if self.verbose:
            print("🔧 Generating Excel workbook...")
        
        # Create all sheets
        self.create_summary_sheet()
        self.create_collaboration_matrix_sheet()
        self.create_collaboration_scores_sheet()
        self.create_temporal_analysis_sheet()
        
        if self.insights_data:
            self.create_ml_insights_sheet()
        
        self.create_formulas_sheet()
        
        # Add charts and named ranges
        self.create_charts()
        self.add_named_ranges()
        
        # Save workbook
        return self.save_workbook(output_path)

def main():
    """Main execution function."""
    parser = argparse.ArgumentParser(
        description='Developer Collaboration Matrix - Excel Report Generator',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python main.excel.py --input ./reports/data.json --output ./reports/collaboration_report.xlsx
  python main.excel.py --input data.csv --output report.xlsx --insights ml_insights.json
  python main.excel.py --input report.json --output analysis.xlsx --verbose
        """
    )
    
    parser.add_argument(
        '--input', '-i',
        required=True,
        help='Input JSON or CSV file path'
    )
    
    parser.add_argument(
        '--output', '-o',
        required=True,
        help='Output Excel file path (.xlsx)'
    )
    
    parser.add_argument(
        '--insights',
        help='ML insights JSON file (optional)'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable verbose output'
    )
    
    args = parser.parse_args()
    
    try:
        # Validate output extension
        if not args.output.endswith('.xlsx'):
            args.output += '.xlsx'
        
        # Initialize generator
        generator = CollaborationExcelGenerator(verbose=args.verbose)
        
        if args.verbose:
            print("🚀 Starting Excel report generation...")
            print(f"📄 Input: {args.input}")
            print(f"💾 Output: {args.output}")
            if args.insights:
                print(f"🧠 Insights: {args.insights}")
        
        # Load data
        generator.load_data(args.input, args.insights)
        
        # Generate report
        output_file = generator.generate_report(args.output)
        
        # Summary
        print(f"\n📊 EXCEL REPORT SUMMARY:")
        print(f"📄 Input file: {args.input}")
        if args.insights:
            print(f"🧠 Insights file: {args.insights}")
        print(f"💾 Output file: {output_file}")
        print(f"📋 Sheets created: {len(generator.workbook.sheetnames)}")
        print(f"📊 Sheet names: {', '.join(generator.workbook.sheetnames)}")
        print("\n✅ Excel report generated successfully!")
        
    except FileNotFoundError as e:
        print(f"❌ File not found: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"❌ Data error: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()