#!/usr/bin/env python3
"""
Excel Report Generator for Merge Readiness & Quality Score Analysis
Creates interactive Excel workbooks with multiple sheets and formulas
"""

import json
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import click
import os
from datetime import datetime
import sys
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
import numpy as np
from rich.console import Console
from rich.markup import escape

# Safe console printing utility
console = Console()

def safe_console_print(message: str, style: str = "") -> None:
    """Safely print messages to console with Rich formatting."""
    escaped = escape(message)
    try:
        if style:
            console.print(f"[{style}]{escaped}[/{style}]")
        else:
            console.print(escaped)
    except Exception as e:
        console.print(f"[red]⚠ Print error:[/red] {escape(str(e))}")
        console.print(escaped)

class TypeSafeExtractor:
    """Utility class for safely extracting and converting data types."""
    
    @staticmethod
    def safe_int(value: Any, default: int = 0) -> int:
        """Safely convert value to int."""
        if pd.isna(value) or value is None:
            return default
        try:
            if isinstance(value, (np.integer, np.int64, np.int32)):
                return int(value)
            return int(float(value))
        except (ValueError, TypeError, OverflowError):
            return default
    
    @staticmethod
    def safe_float(value: Any, default: float = 0.0) -> float:
        """Safely convert value to float."""
        if pd.isna(value) or value is None:
            return default
        try:
            if isinstance(value, (np.floating, np.float64, np.float32)):
                return float(value)
            return float(value)
        except (ValueError, TypeError, OverflowError):
            return default
    
    @staticmethod
    def safe_str(value: Any, default: str = "") -> str:
        """Safely convert value to string."""
        if pd.isna(value) or value is None:
            return default
        try:
            return str(value)
        except Exception:
            return default
    
    @staticmethod
    def safe_dict_get(data: Dict[str, Any], key: str, default: Any = None) -> Any:
        """Safely get value from dictionary with type checking."""
        try:
            value = data.get(key, default)
            return value if value is not None else default
        except Exception:
            return default

class MergeReadinessExcelGenerator:
    """Generate Excel reports for merge readiness and quality analysis."""
    
    def __init__(self):
        self.workbook = Workbook()
        self.json_data: Dict[str, Any] = {}
        self.csv_data: Optional[pd.DataFrame] = None
        self.extractor = TypeSafeExtractor()
        self.setup_styles()
        
    def setup_styles(self) -> None:
        """Define consistent styling for the workbook."""
        self.styles = {
            'header': Font(bold=True, size=14, color='FFFFFF'),
            'subheader': Font(bold=True, size=12, color='2E86AB'),
            'data': Font(size=10),
            'metric': Font(bold=True, size=11),
            'header_fill': PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid'),
            'alt_fill': PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center')
        }
    
    def load_data(self, json_file: str, csv_file: Optional[str] = None) -> None:
        """Load data from JSON and optionally CSV files."""
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                self.json_data = json.load(f)
            
            safe_console_print(f"✅ Loaded JSON data from {json_file}", "green")
            
            if csv_file and os.path.exists(csv_file):
                self.csv_data = pd.read_csv(csv_file)
                safe_console_print(f"✅ Loaded CSV data from {csv_file}", "green")
            else:
                self.csv_data = None
                
        except FileNotFoundError as e:
            raise Exception(f"File not found: {str(e)}")
        except json.JSONDecodeError as e:
            raise Exception(f"Invalid JSON format: {str(e)}")
        except Exception as e:
            raise Exception(f"Failed to load data: {str(e)}")
    
    def safe_set_cell_value(self, worksheet, cell_ref: str, value: Any) -> None:
        """Safely set cell value with proper type conversion."""
        try:
            if isinstance(value, str) and value.startswith('='):
                # It's a formula
                worksheet[cell_ref] = value
            elif isinstance(value, (int, float)):
                # Numeric value
                worksheet[cell_ref] = value
            elif isinstance(value, (np.integer, np.int64, np.int32)):
                # NumPy integer
                worksheet[cell_ref] = int(value)
            elif isinstance(value, (np.floating, np.float64, np.float32)):
                # NumPy float
                worksheet[cell_ref] = float(value)
            elif pd.isna(value) or value is None:
                # Handle NaN or None
                worksheet[cell_ref] = 0
            else:
                # Convert to string as fallback
                worksheet[cell_ref] = str(value)
        except Exception as e:
            safe_console_print(f"Warning: Could not set cell {cell_ref} with value {value}: {str(e)}", "yellow")
            worksheet[cell_ref] = str(value) if value is not None else ""
    
    def create_summary_sheet(self) -> None:
        """Create executive summary sheet."""
        # Remove default sheet and create summary
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        ws = self.workbook.create_sheet('Executive Summary', 0)
        
        # Title
        ws['A1'] = 'Merge Readiness & Quality Score Analysis'
        ws['A1'].font = Font(bold=True, size=18, color='2E86AB')
        ws.merge_cells('A1:F1')
        
        # Date range
        date_range = self.extractor.safe_dict_get(self.json_data, 'date_range', {})
        start_date = self.extractor.safe_dict_get(date_range, 'start_date', 'N/A')
        end_date = self.extractor.safe_dict_get(date_range, 'end_date', 'N/A')
        
        ws['A3'] = f"Analysis Period: {start_date} to {end_date}"
        ws['A3'].font = self.styles['subheader']
        ws.merge_cells('A3:F3')
        
        # Summary metrics
        summary = self.extractor.safe_dict_get(self.json_data, 'summary', {})
        
        # Create summary table
        headers = ['Metric', 'Value', 'Status']
        ws.append([''] * 6)  # Empty row
        ws.append(headers + [''] * 3)
        
        # Style header row
        header_row = ws.max_row
        for col in range(1, 4):
            cell = ws.cell(row=header_row, column=col)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['border']
            cell.alignment = self.styles['center']
        
        # Add summary data with safe extraction
        summary_data = [
            ['Total Repositories', self.extractor.safe_int(summary.get('total_repositories', 0))],
            ['Total Pull Requests', self.extractor.safe_int(summary.get('total_pull_requests', 0))],
            ['Linked Issue-PR Pairs', self.extractor.safe_int(summary.get('linked_issue_pr_pairs', 0))],
            ['Avg Lead Time (hours)', self.extractor.safe_float(summary.get('avg_lead_time_hours', 0))],
            ['Median Lead Time (hours)', self.extractor.safe_float(summary.get('median_lead_time_hours', 0))],
            ['Merge Readiness Score', self.extractor.safe_float(summary.get('merge_readiness_score', 0))],
            ['Quality Score', self.extractor.safe_float(summary.get('quality_score', 0))],
            ['Bottlenecks Detected', self.extractor.safe_int(summary.get('bottlenecks_detected', 0))]
        ]
        
        for row_data in summary_data:
            current_row = ws.max_row + 1
            self.safe_set_cell_value(ws, f'A{current_row}', row_data[0])
            self.safe_set_cell_value(ws, f'B{current_row}', row_data[1])
            self.safe_set_cell_value(ws, f'C{current_row}', self.get_status_indicator(row_data[1], row_data[0]))
            
            # Style data cells
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.border = self.styles['border']
                cell.alignment = self.styles['left'] if col == 1 else self.styles['center']
                if col == 2:
                    cell.font = self.styles['metric']
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        # Add conditional formatting safely
        try:
            self.add_conditional_formatting(ws, f'B{header_row+1}:B{header_row+8}', 'score')
        except Exception as e:
            safe_console_print(f"Warning: Could not add conditional formatting: {str(e)}", "yellow")
    
    def create_lead_time_analysis_sheet(self) -> None:
        """Create detailed lead time analysis sheet."""
        ws = self.workbook.create_sheet('Lead Time Analysis')
        
        # Title
        ws['A1'] = 'Lead Time Analysis'
        ws['A1'].font = Font(bold=True, size=16, color='2E86AB')
        ws.merge_cells('A1:E1')
        
        detailed_analysis = self.extractor.safe_dict_get(self.json_data, 'detailed_analysis', {})
        lead_metrics = self.extractor.safe_dict_get(detailed_analysis, 'lead_time_metrics', {})
        
        # Lead time statistics table
        ws['A3'] = 'Lead Time Statistics'
        ws['A3'].font = self.styles['subheader']
        
        headers = ['Metric', 'Value (Hours)', 'Value (Days)', 'Percentile']
        ws.append([''] * 6)  # Empty row
        ws.append(headers + [''] * 2)
        
        # Style headers
        header_row = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=header_row, column=col)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['border']
            cell.alignment = self.styles['center']
        
        # Add lead time data
        lead_data = [
            ['Total Pairs', self.extractor.safe_int(lead_metrics.get('total_pairs', 0)), 'N/A', 'N/A'],
            ['Average Lead Time', self.extractor.safe_float(lead_metrics.get('avg_lead_time_hours', 0))],
            ['Median Lead Time', self.extractor.safe_float(lead_metrics.get('median_lead_time_hours', 0))],
            ['75th Percentile', self.extractor.safe_float(lead_metrics.get('p75_lead_time_hours', 0))],
            ['95th Percentile', self.extractor.safe_float(lead_metrics.get('p95_lead_time_hours', 0))],
            ['Minimum Lead Time', self.extractor.safe_float(lead_metrics.get('min_lead_time_hours', 0))],
            ['Maximum Lead Time', self.extractor.safe_float(lead_metrics.get('max_lead_time_hours', 0))]
        ]
        
        for i, row_data in enumerate(lead_data):
            current_row = ws.max_row + 1
            self.safe_set_cell_value(ws, f'A{current_row}', row_data[0])
            self.safe_set_cell_value(ws, f'B{current_row}', row_data[1])
            
            if i == 0:  # First row (total pairs)
                self.safe_set_cell_value(ws, f'C{current_row}', row_data[2])
                self.safe_set_cell_value(ws, f'D{current_row}', row_data[3])
            else:
                # Calculate days
                days_value = self.extractor.safe_float(row_data[1]) / 24
                self.safe_set_cell_value(ws, f'C{current_row}', round(days_value, 2))
                
                percentile_labels = ['Mean', '50th', '75th', '95th', 'Min', 'Max']
                self.safe_set_cell_value(ws, f'D{current_row}', percentile_labels[i-1])
            
            # Style cells
            for col in range(1, 5):
                cell = ws.cell(row=current_row, column=col)
                cell.border = self.styles['border']
                cell.alignment = self.styles['left'] if col == 1 else self.styles['center']
                if col == 2 and isinstance(row_data[1], (int, float)):
                    cell.font = self.styles['metric']
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
    
    def create_quality_metrics_sheet(self) -> None:
        """Create quality metrics analysis sheet."""
        ws = self.workbook.create_sheet('Quality Metrics')
        
        # Title
        ws['A1'] = 'Quality Metrics Analysis'
        ws['A1'].font = Font(bold=True, size=16, color='2E86AB')
        ws.merge_cells('A1:E1')
        
        detailed_analysis = self.extractor.safe_dict_get(self.json_data, 'detailed_analysis', {})
        quality_metrics = self.extractor.safe_dict_get(detailed_analysis, 'quality_metrics', {})
        
        # Quality overview table
        ws['A3'] = 'Quality Overview'
        ws['A3'].font = self.styles['subheader']
        
        headers = ['Metric', 'Value', 'Target', 'Performance']
        ws.append([''] * 6)  # Empty row
        ws.append(headers + [''] * 2)
        
        # Style headers
        header_row = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=header_row, column=col)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['border']
            cell.alignment = self.styles['center']
        
        # Add quality data with safe extraction
        quality_data = [
            ['Overall Score', self.extractor.safe_float(quality_metrics.get('overall_score', 0)), 85],
            ['Total PRs', self.extractor.safe_int(quality_metrics.get('total_prs', 0)), 'N/A'],
            ['Merged PRs', self.extractor.safe_int(quality_metrics.get('merged_prs', 0)), 'N/A'],
            ['Reverted PRs', self.extractor.safe_int(quality_metrics.get('reverted_prs', 0)), 0],
            ['Merge Success Rate', self.extractor.safe_float(quality_metrics.get('merge_success_rate', 0)), 95],
            ['Avg Comments per PR', self.extractor.safe_float(quality_metrics.get('avg_comments_per_pr', 0)), 3],
            ['Comment to LOC Ratio', self.extractor.safe_float(quality_metrics.get('comment_to_loc_ratio', 0)), 0.01]
        ]
        
        for row_data in quality_data:
            current_row = ws.max_row + 1
            self.safe_set_cell_value(ws, f'A{current_row}', row_data[0])
            self.safe_set_cell_value(ws, f'B{current_row}', row_data[1])
            self.safe_set_cell_value(ws, f'C{current_row}', row_data[2])
            
            # Calculate performance indicator
            if isinstance(row_data[2], (int, float)) and row_data[2] != 'N/A':
                if row_data[1] >= row_data[2]:
                    performance = "Exceeds" if row_data[0] != 'Reverted PRs' else "Good"
                else:
                    performance = "Below" if row_data[0] != 'Reverted PRs' else "Review Needed"
            else:
                performance = "Count"
            
            self.safe_set_cell_value(ws, f'D{current_row}', performance)
            
            # Style cells
            for col in range(1, 5):
                cell = ws.cell(row=current_row, column=col)
                cell.border = self.styles['border']
                cell.alignment = self.styles['left'] if col == 1 else self.styles['center']
                if col == 2 and isinstance(row_data[1], (int, float)):
                    cell.font = self.styles['metric']
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def create_repository_breakdown_sheet(self) -> None:
        """Create repository-level breakdown sheet."""
        ws = self.workbook.create_sheet('Repository Breakdown')
        
        # Title
        ws['A1'] = 'Repository-Level Analysis'
        ws['A1'].font = Font(bold=True, size=16, color='2E86AB')
        ws.merge_cells('A1:G1')
        
        detailed_analysis = self.extractor.safe_dict_get(self.json_data, 'detailed_analysis', {})
        repo_breakdown = self.extractor.safe_dict_get(detailed_analysis, 'repository_breakdown', [])
        
        if repo_breakdown:
            # Create repository table
            headers = ['Repository', 'Issues', 'PRs', 'Linked Pairs', 'Avg Lead Time (hrs)', 'Quality Score', 'Performance']
            ws.append([''] * 8)  # Empty row
            ws.append(headers + [''])
            
            # Style headers
            header_row = ws.max_row
            for col in range(1, 8):
                cell = ws.cell(row=header_row, column=col)
                cell.font = self.styles['header']
                cell.fill = self.styles['header_fill']
                cell.border = self.styles['border']
                cell.alignment = self.styles['center']
            
            # Add repository data
            for repo in repo_breakdown:
                current_row = ws.max_row + 1
                
                repo_name = self.extractor.safe_str(repo.get('repository', ''), 'Unknown')
                issues_count = self.extractor.safe_int(repo.get('issues_count', 0))
                prs_count = self.extractor.safe_int(repo.get('prs_count', 0))
                linked_pairs = self.extractor.safe_int(repo.get('linked_pairs_count', 0))
                avg_lead_time = self.extractor.safe_float(repo.get('avg_lead_time_hours', 0))
                quality_score = self.extractor.safe_float(repo.get('quality_score', 0))
                
                # Determine performance
                if avg_lead_time <= 24:
                    performance = "Excellent"
                elif avg_lead_time <= 72:
                    performance = "Good"
                else:
                    performance = "Needs Work"
                
                self.safe_set_cell_value(ws, f'A{current_row}', repo_name)
                self.safe_set_cell_value(ws, f'B{current_row}', issues_count)
                self.safe_set_cell_value(ws, f'C{current_row}', prs_count)
                self.safe_set_cell_value(ws, f'D{current_row}', linked_pairs)
                self.safe_set_cell_value(ws, f'E{current_row}', avg_lead_time)
                self.safe_set_cell_value(ws, f'F{current_row}', quality_score)
                self.safe_set_cell_value(ws, f'G{current_row}', performance)
                
                # Style cells
                for col in range(1, 8):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = self.styles['border']
                    cell.alignment = self.styles['left'] if col == 1 else self.styles['center']
                    if col in [2, 3, 4, 5, 6]:
                        cell.font = self.styles['metric']
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 12
        ws.column_dimensions['G'].width = 15
    
    def get_status_indicator(self, value: Union[int, float], metric_name: str) -> str:
        """Get status indicator based on metric type and value."""
        if 'Lead Time' in metric_name:
            if value <= 24:
                return '🟢'
            elif value <= 72:
                return '🟡'
            else:
                return '🔴'
        elif 'Score' in metric_name:
            if value >= 85:
                return '🟢'
            elif value >= 70:
                return '🟡'
            else:
                return '🔴'
        elif 'Bottleneck' in metric_name:
            return '🔴' if value > 0 else '🟢'
        else:
            return '📊' if value > 0 else '❌'
    
    def add_conditional_formatting(self, worksheet, cell_range: str, format_type: str) -> None:
        """Add conditional formatting to specified range."""
        try:
            if format_type == 'score':
                # Color scale for scores (red to green)
                rule = ColorScaleRule(
                    start_type='num', start_value=0, start_color='FF6B6B',
                    mid_type='num', mid_value=50, mid_color='FFE66D',
                    end_type='num', end_value=100, end_color='4ECDC4'
                )
                worksheet.conditional_formatting.add(cell_range, rule)
            
            elif format_type == 'lead_time':
                # Color scale for lead times (green to red)
                rule = ColorScaleRule(
                    start_type='num', start_value=0, start_color='4ECDC4',
                    mid_type='num', mid_value=72, mid_color='FFE66D',
                    end_type='num', end_value=168, end_color='FF6B6B'
                )
                worksheet.conditional_formatting.add(cell_range, rule)
            
            elif format_type == 'quality':
                # Data bars for quality metrics
                rule = DataBarRule(
                    start_type='min', start_value=None,
                    end_type='max', end_value=None,
                    color='4ECDC4'
                )
                worksheet.conditional_formatting.add(cell_range, rule)
        except Exception as e:
            safe_console_print(f"Warning: Could not add conditional formatting for {cell_range}: {str(e)}", "yellow")
    
    def generate_excel_report(self, output_file: str) -> str:
        """Generate the complete Excel report."""
        try:
            safe_console_print("📊 Creating Excel sheets...", "blue")
            
            # Create all sheets
            self.create_summary_sheet()
            self.create_lead_time_analysis_sheet()
            self.create_quality_metrics_sheet()
            self.create_repository_breakdown_sheet()
            
            safe_console_print("💾 Saving Excel workbook...", "blue")
            
            # Ensure output directory exists
            output_path = Path(output_file)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            self.workbook.save(output_file)
            
            safe_console_print(f"✅ Excel report saved to: {output_file}", "green")
            return output_file
            
        except Exception as e:
            raise Exception(f"Failed to generate Excel report: {str(e)}")

@click.command()
@click.option('--json', 'json_file', required=True, help='Input JSON report file', type=click.Path(exists=True))
@click.option('--csv', 'csv_file', help='Optional CSV data file', type=click.Path(exists=True))
@click.option('-o', '--output', required=True, help='Output Excel file path')
@click.option('-v', '--verbose', is_flag=True, help='Verbose output')
def cli(json_file: str, csv_file: Optional[str], output: str, verbose: bool) -> None:
    """Generate Excel Report for Merge Readiness Analysis."""
    try:
        # Initialize generator
        generator = MergeReadinessExcelGenerator()
        
        # Load data
        if verbose:
            safe_console_print("📥 Loading data...", "blue")
        generator.load_data(json_file, csv_file)
        
        # Generate report
        generator.generate_excel_report(output)
        
        if verbose:
            safe_console_print("🎯 Excel report generation completed successfully!", "green")
            
    except Exception as e:
        safe_console_print(f"❌ Error: {str(e)}", "red")
        sys.exit(1)

def main() -> None:
    """Main entry point for legacy compatibility."""
    cli()

if __name__ == "__main__":
    cli()