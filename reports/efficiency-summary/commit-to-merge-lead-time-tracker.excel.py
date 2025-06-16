#!/usr/bin/env python3
"""
Commit-to-Merge Lead Time Excel Generator

Creates interactive Excel workbooks with multiple sheets, formulas, and visualizations
from JSON/CSV report data.
"""

import json
import pandas as pd
import argparse
from pathlib import Path
import sys
from datetime import datetime
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataSeries
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import warnings

# Suppress warnings
warnings.filterwarnings('ignore')

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class CommitMergeLeadTimeExcelGenerator:
    """Generate interactive Excel workbooks from lead time analysis data"""
    
    def __init__(self):
        self.workbook = Workbook()
        self.data = None
        self.insights_data = None
        self.setup_styles()
    
    def setup_styles(self):
        """Define styles for the workbook"""
        self.styles = {
            'header': Font(bold=True, color='FFFFFF'),
            'header_fill': PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid'),
            'subheader': Font(bold=True, color='2D3436'),
            'subheader_fill': PatternFill(start_color='DDD6FE', end_color='DDD6FE', fill_type='solid'),
            'metric_value': Font(bold=True, size=14, color='2E86AB'),
            'metric_label': Font(color='6B7280'),
            'border': Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            ),
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center')
        }
    
    def load_data(self, json_path: str, insights_path: str = None):
        """Load data from JSON files"""
        try:
            with open(json_path, 'r') as f:
                self.data = json.load(f)
            
            if insights_path and Path(insights_path).exists():
                with open(insights_path, 'r') as f:
                    self.insights_data = json.load(f)
            
            logger.info(f"Data loaded successfully from {json_path}")
            
        except Exception as e:
            logger.error(f"Error loading data: {str(e)}")
            raise
    
    def create_summary_sheet(self):
        """Create the executive summary sheet"""
        # Remove default sheet if it exists
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        ws = self.workbook.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = 'Commit-to-Merge Lead Time Analysis'
        ws['A1'].font = Font(bold=True, size=18, color='2E86AB')
        ws.merge_cells('A1:F1')
        
        # Analysis date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, color='6B7280')
        
        # Date range
        date_range = self.data.get('date_range', {})
        if date_range.get('start_date') and date_range.get('end_date'):
            ws['A3'] = f"Analysis Period: {date_range['start_date']} to {date_range['end_date']}"
        else:
            ws['A3'] = "Analysis Period: Complete repository history"
        ws['A3'].font = Font(size=10, color='6B7280')
        
        # Key metrics section
        ws['A5'] = 'Key Metrics'
        ws['A5'].font = self.styles['subheader']
        ws['A5'].fill = self.styles['subheader_fill']
        ws.merge_cells('A5:B5')
        
        summary = self.data.get('summary', {})
        total = self.data.get('total', {})
        
        # Metrics with formulas
        metrics_data = [
            ('Total Pull Requests', summary.get('TOTAL_PRS', 0), 'COUNT of all analyzed PRs'),
            ('Merged Pull Requests', summary.get('MERGED_PRS', 0), 'COUNT of successfully merged PRs'),
            ('Average Lead Time (Days)', '=ROUND(Summary_Data!C6/24,2)', 'MEAN of all lead times'),
            ('Median Lead Time (Days)', '=ROUND(Summary_Data!C7/24,2)', 'MEDIAN of all lead times'),
            ('Fastest PR (Days)', '=ROUND(Summary_Data!C8/24,2)', 'MIN lead time'),
            ('Slowest PR (Days)', '=ROUND(Summary_Data!C9/24,2)', 'MAX lead time'),
            ('Total Contributors', total.get('CONTRIBUTORS', 0), 'DISTINCT count of authors'),
            ('Repositories Analyzed', total.get('REPOSITORIES_ANALYZED', 0), 'DISTINCT count of repositories')
        ]
        
        row = 6
        for metric, value, description in metrics_data:
            ws[f'A{row}'] = metric
            ws[f'A{row}'].font = self.styles['metric_label']
            
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = self.styles['metric_value']
            ws[f'B{row}'].alignment = self.styles['center']
            
            ws[f'C{row}'] = description
            ws[f'C{row}'].font = Font(size=9, color='9CA3AF')
            
            # Add borders
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = self.styles['border']
            
            row += 1
        
        # Performance indicators
        ws['A15'] = 'Performance Assessment'
        ws['A15'].font = self.styles['subheader']
        ws['A15'].fill = self.styles['subheader_fill']
        ws.merge_cells('A15:C15')
        
        # Speed assessment
        avg_days = summary.get('AVG_LEAD_TIME_DAYS', 0)
        if avg_days > 7:
            speed_assessment = 'Needs Improvement'
            speed_color = 'C9184A'
        elif avg_days > 3:
            speed_assessment = 'Fair'
            speed_color = 'F77F00'
        else:
            speed_assessment = 'Good'
            speed_color = '6A994E'
        
        ws['A16'] = 'Delivery Speed'
        ws['B16'] = speed_assessment
        ws['B16'].font = Font(bold=True, color=speed_color)
        
        # Consistency assessment
        variance = summary.get('MAX_LEAD_TIME_DAYS', 0) - summary.get('MIN_LEAD_TIME_DAYS', 0)
        if variance < 7:
            consistency = 'Consistent'
            consistency_color = '6A994E'
        elif variance < 14:
            consistency = 'Moderate'
            consistency_color = 'F77F00'
        else:
            consistency = 'Variable'
            consistency_color = 'C9184A'
        
        ws['A17'] = 'Process Consistency'
        ws['B17'] = consistency
        ws['B17'].font = Font(bold=True, color=consistency_color)
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        
        return ws
    
    def create_summary_data_sheet(self):
        """Create a hidden sheet with summary data for formulas"""
        ws = self.workbook.create_sheet("Summary_Data")
        ws.sheet_state = 'hidden'
        
        summary = self.data.get('summary', {})
        
        # Headers
        ws['A1'] = 'Metric'
        ws['B1'] = 'Value_Hours'
        ws['C1'] = 'Value_Days'
        
        # Data for formulas
        data = [
            ('TOTAL_PRS', summary.get('TOTAL_PRS', 0), summary.get('TOTAL_PRS', 0)),
            ('MERGED_PRS', summary.get('MERGED_PRS', 0), summary.get('MERGED_PRS', 0)),
            ('AVG_LEAD_TIME', summary.get('AVG_LEAD_TIME_HOURS', 0), summary.get('AVG_LEAD_TIME_DAYS', 0)),
            ('MEDIAN_LEAD_TIME', summary.get('MEDIAN_LEAD_TIME_HOURS', 0), summary.get('MEDIAN_LEAD_TIME_DAYS', 0)),
            ('MIN_LEAD_TIME', summary.get('MIN_LEAD_TIME_DAYS', 0) * 24, summary.get('MIN_LEAD_TIME_DAYS', 0)),
            ('MAX_LEAD_TIME', summary.get('MAX_LEAD_TIME_DAYS', 0) * 24, summary.get('MAX_LEAD_TIME_DAYS', 0)),
            ('P75_LEAD_TIME', summary.get('P75_LEAD_TIME_DAYS', 0) * 24, summary.get('P75_LEAD_TIME_DAYS', 0)),
            ('P95_LEAD_TIME', summary.get('P95_LEAD_TIME_DAYS', 0) * 24, summary.get('P95_LEAD_TIME_DAYS', 0))
        ]
        
        for i, (metric, hours_val, days_val) in enumerate(data, 2):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = hours_val
            ws[f'C{i}'] = days_val
        
        return ws
    
    def create_detailed_data_sheet(self):
        """Create detailed pull request data sheet"""
        ws = self.workbook.create_sheet("Pull Request Details")
        
        # Get pull request data
        prs = self.data.get('detailed_analysis', {}).get('pull_requests', [])
        
        if not prs:
            ws['A1'] = 'No pull request data available'
            return ws
        
        # Convert to DataFrame for easier handling
        df = pd.DataFrame(prs)
        
        # Select and order columns
        columns = ['pr_number', 'title', 'author', 'repository', 'state', 
                  'first_commit_timestamp', 'merge_timestamp', 
                  'LEAD_TIME_HOURS', 'LEAD_TIME_DAYS', 'head_branch', 'base_branch']
        
        available_columns = [col for col in columns if col in df.columns]
        df_filtered = df[available_columns].copy()
        
        # Clean data for Excel
        for col in df_filtered.columns:
            if df_filtered[col].dtype == 'object':
                df_filtered[col] = df_filtered[col].astype(str)
        
        # Add header row
        headers = {
            'pr_number': 'PR Number',
            'title': 'Title',
            'author': 'Author',
            'repository': 'Repository',
            'state': 'State',
            'first_commit_timestamp': 'First Commit',
            'merge_timestamp': 'Merge Time',
            'LEAD_TIME_HOURS': 'Lead Time (Hours)',
            'LEAD_TIME_DAYS': 'Lead Time (Days)',
            'head_branch': 'Source Branch',
            'base_branch': 'Target Branch'
        }
        
        # Write headers
        for i, col in enumerate(df_filtered.columns, 1):
            cell = ws.cell(row=1, column=i)
            cell.value = headers.get(col, col)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.alignment = self.styles['center']
            cell.border = self.styles['border']
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df_filtered, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.border = self.styles['border']
                
                # Special formatting for numeric columns
                if c_idx in [8, 9]:  # Lead time columns
                    cell.alignment = self.styles['center']
                    if isinstance(value, (int, float)) and value > 0:
                        cell.number_format = '0.00'
        
        # Add conditional formatting for lead times
        lead_time_col = None
        for i, col in enumerate(df_filtered.columns, 1):
            if col == 'LEAD_TIME_DAYS':
                lead_time_col = get_column_letter(i)
                break
        
        if lead_time_col:
            # Color scale for lead times
            color_scale = ColorScaleRule(
                start_type='min', start_color='63BE7B',
                mid_type='percentile', mid_value=50, mid_color='FFDD00',
                end_type='max', end_color='F87979'
            )
            ws.conditional_formatting.add(f'{lead_time_col}2:{lead_time_col}{len(df_filtered) + 1}', color_scale)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return ws
    
    def create_trends_sheet(self):
        """Create trends analysis sheet with charts"""
        ws = self.workbook.create_sheet("Trends Analysis")
        
        trends = self.data.get('detailed_analysis', {}).get('trends', {})
        
        # Title
        ws['A1'] = 'Lead Time Trends Analysis'
        ws['A1'].font = Font(bold=True, size=16, color='2E86AB')
        ws.merge_cells('A1:E1')
        
        if not trends:
            ws['A3'] = 'No trend data available'
            return ws
        
        # Weekly trends
        weekly_data = trends.get('weekly', {})
        if weekly_data:
            ws['A3'] = 'Weekly Trends'
            ws['A3'].font = self.styles['subheader']
            ws['A3'].fill = self.styles['subheader_fill']
            
            # Headers
            headers = ['Week', 'PR Count', 'Avg Lead Time (Days)', 'Median Lead Time (Days)', 'Min', 'Max']
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=i)
                cell.value = header
                cell.font = self.styles['header']
                cell.fill = self.styles['header_fill']
                cell.border = self.styles['border']
            
            # Data
            sorted_weeks = sorted(weekly_data.items())
            for i, (week, data) in enumerate(sorted_weeks, 5):
                ws[f'A{i}'] = week
                ws[f'B{i}'] = data.get('PR_COUNT', 0)
                ws[f'C{i}'] = round(data.get('MEAN', 0), 2)
                ws[f'D{i}'] = round(data.get('MEDIAN', 0), 2)
                ws[f'E{i}'] = round(data.get('MIN', 0), 2)
                ws[f'F{i}'] = round(data.get('MAX', 0), 2)
                
                # Add borders
                for col in range(1, 7):
                    ws.cell(row=i, column=col).border = self.styles['border']
            
            # Create chart
            if len(sorted_weeks) > 1:
                chart = LineChart()
                chart.title = "Weekly Lead Time Trends"
                chart.style = 10
                chart.x_axis.title = 'Week'
                chart.y_axis.title = 'Lead Time (Days)'
                
                # Data for chart
                data_ref = Reference(ws, min_col=3, min_row=4, max_col=4, max_row=4 + len(sorted_weeks))
                categories = Reference(ws, min_col=1, min_row=5, max_row=4 + len(sorted_weeks))
                
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(categories)
                
                ws.add_chart(chart, "H4")
        
        # Monthly trends
        monthly_data = trends.get('monthly', {})
        if monthly_data:
            start_row = 15 if weekly_data else 3
            
            ws[f'A{start_row}'] = 'Monthly Trends'
            ws[f'A{start_row}'].font = self.styles['subheader']
            ws[f'A{start_row}'].fill = self.styles['subheader_fill']
            
            # Headers
            header_row = start_row + 1
            headers = ['Month', 'PR Count', 'Avg Lead Time (Days)', 'Median Lead Time (Days)', 'Min', 'Max']
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=i)
                cell.value = header
                cell.font = self.styles['header']
                cell.fill = self.styles['header_fill']
                cell.border = self.styles['border']
            
            # Data
            sorted_months = sorted(monthly_data.items())
            for i, (month, data) in enumerate(sorted_months, header_row + 1):
                ws[f'A{i}'] = month
                ws[f'B{i}'] = data.get('PR_COUNT', 0)
                ws[f'C{i}'] = round(data.get('MEAN', 0), 2)
                ws[f'D{i}'] = round(data.get('MEDIAN', 0), 2)
                ws[f'E{i}'] = round(data.get('MIN', 0), 2)
                ws[f'F{i}'] = round(data.get('MAX', 0), 2)
                
                # Add borders
                for col in range(1, 7):
                    ws.cell(row=i, column=col).border = self.styles['border']
            
            # Create monthly chart
            if len(sorted_months) > 1:
                chart = LineChart()
                chart.title = "Monthly Lead Time Trends"
                chart.style = 12
                chart.x_axis.title = 'Month'
                chart.y_axis.title = 'Lead Time (Days)'
                
                # Data for chart
                data_ref = Reference(ws, min_col=3, min_row=header_row, max_col=4, max_row=header_row + len(sorted_months))
                categories = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + len(sorted_months))
                
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(categories)
                
                chart_position = f"H{start_row + 2}"
                ws.add_chart(chart, chart_position)
        
        return ws
    
    def create_contributors_sheet(self):
        """Create contributor analysis sheet"""
        ws = self.workbook.create_sheet("Contributors Analysis")
        
        contributors = self.data.get('detailed_analysis', {}).get('contributor_metrics', {})
        
        # Title
        ws['A1'] = 'Contributor Performance Analysis'
        ws['A1'].font = Font(bold=True, size=16, color='2E86AB')
        ws.merge_cells('A1:G1')
        
        if not contributors:
            ws['A3'] = 'No contributor data available'
            return ws
        
        # Headers
        headers = ['Contributor', 'Total PRs', 'Avg Lead Time (Days)', 'Median Lead Time', 'Min Lead Time', 'Max Lead Time', 'Std Dev']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i)
            cell.value = header
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['border']
        
        # Sort contributors by PR count
        sorted_contributors = sorted(contributors.items(), key=lambda x: x[1].get('TOTAL_PRS', 0), reverse=True)
        
        # Data
        for i, (contributor, data) in enumerate(sorted_contributors, 4):
            ws[f'A{i}'] = contributor
            ws[f'B{i}'] = data.get('TOTAL_PRS', 0)
            ws[f'C{i}'] = round(data.get('MEAN', 0), 2)
            ws[f'D{i}'] = round(data.get('MEDIAN', 0), 2)
            ws[f'E{i}'] = round(data.get('MIN', 0), 2)
            ws[f'F{i}'] = round(data.get('MAX', 0), 2)
            ws[f'G{i}'] = round(data.get('std', 0), 2)
            
            # Add borders
            for col in range(1, 8):
                ws.cell(row=i, column=col).border = self.styles['border']
        
        # Create top contributors chart
        if len(sorted_contributors) > 1:
            chart = BarChart()
            chart.title = "Top Contributors by PR Count"
            chart.style = 10
            chart.x_axis.title = 'Contributor'
            chart.y_axis.title = 'Number of PRs'
            
            # Limit to top 10 contributors for readability
            max_contributors = min(10, len(sorted_contributors))
            
            data_ref = Reference(ws, min_col=2, min_row=3, max_row=3 + max_contributors)
            categories = Reference(ws, min_col=1, min_row=4, max_row=3 + max_contributors)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(categories)
            
            ws.add_chart(chart, "I4")
        
        # Conditional formatting for lead times
        if len(sorted_contributors) > 0:
            end_row = 3 + len(sorted_contributors)
            
            # Highlight fast contributors (green)
            fast_rule = CellIsRule(operator='lessThan', formula=[3], stopIfTrue=True, 
                                  fill=PatternFill(start_color='C6F6D5', end_color='C6F6D5', fill_type='solid'))
            ws.conditional_formatting.add(f'C4:C{end_row}', fast_rule)
            
            # Highlight slow contributors (red)
            slow_rule = CellIsRule(operator='greaterThan', formula=[7], stopIfTrue=True,
                                  fill=PatternFill(start_color='FED7D7', end_color='FED7D7', fill_type='solid'))
            ws.conditional_formatting.add(f'C4:C{end_row}', slow_rule)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return ws
    
    def create_ml_insights_sheet(self):
        """Create ML insights sheet if data is available"""
        if not self.insights_data:
            return None
            
        ws = self.workbook.create_sheet("ML Insights")
        
        # Title
        ws['A1'] = 'Machine Learning Analysis Insights'
        ws['A1'].font = Font(bold=True, size=16, color='2E86AB')
        ws.merge_cells('A1:F1')
        
        row = 3
        
        # Clustering Analysis
        clustering = self.insights_data.get('clustering_analysis', {})
        if clustering:
            ws[f'A{row}'] = 'Process Pattern Clusters'
            ws[f'A{row}'].font = self.styles['subheader']
            ws[f'A{row}'].fill = self.styles['subheader_fill']
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            # Headers
            headers = ['Pattern', 'PR Count', 'Avg Lead Time (Days)', 'Description']
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=i)
                cell.value = header
                cell.font = self.styles['header']
                cell.fill = self.styles['header_fill']
                cell.border = self.styles['border']
            row += 1
            
            # Cluster data
            for cluster_id, data in clustering.items():
                ws[f'A{row}'] = cluster_id.replace('_', ' ').title()
                ws[f'B{row}'] = data.get('count', 0)
                ws[f'C{row}'] = round(data.get('avg_lead_time_days', 0), 2)
                ws[f'D{row}'] = data.get('characteristics', '')
                
                # Add borders
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = self.styles['border']
                row += 1
            
            row += 2
        
        # Anomaly Detection
        anomalies = self.insights_data.get('predictive_insights', {}).get('anomalies', {})
        if anomalies:
            ws[f'A{row}'] = 'Anomaly Detection Results'
            ws[f'A{row}'].font = self.styles['subheader']
            ws[f'A{row}'].fill = self.styles['subheader_fill']
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            ws[f'A{row}'] = 'Outlier PRs Detected'
            ws[f'B{row}'] = anomalies.get('count', 0)
            ws[f'B{row}'].font = self.styles['metric_value']
            row += 1
            
            ws[f'A{row}'] = 'Percentage of Total'
            ws[f'B{row}'] = f"{anomalies.get('percentage', 0):.1f}%"
            ws[f'B{row}'].font = self.styles['metric_value']
            row += 1
            
            ws[f'A{row}'] = 'Threshold (Days)'
            ws[f'B{row}'] = round(anomalies.get('threshold_hours', 0) / 24, 1)
            ws[f'B{row}'].font = self.styles['metric_value']
            row += 2
        
        # Recommendations
        recommendations = self.insights_data.get('recommendations', [])
        if recommendations:
            ws[f'A{row}'] = 'AI-Generated Recommendations'
            ws[f'A{row}'].font = self.styles['subheader']
            ws[f'A{row}'].fill = self.styles['subheader_fill']
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            for i, rec in enumerate(recommendations, 1):
                ws[f'A{row}'] = f"{i}."
                ws[f'B{row}'] = rec
                ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        
        return ws
    
    def create_formulas_sheet(self):
        """Create formulas documentation sheet"""
        ws = self.workbook.create_sheet("Formulas Documentation")
        
        # Title
        ws['A1'] = 'Formulas and Calculations Reference'
        ws['A1'].font = Font(bold=True, size=16, color='2E86AB')
        ws.merge_cells('A1:D1')
        
        # Headers
        headers = ['Formula Name', 'Formula', 'Description', 'Example']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i)
            cell.value = header
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['border']
        
        # Formula definitions from the report
        formulas = self.data.get('formulas', {})
        
        # Default formulas if not in data
        if not formulas:
            formulas = {
                'LEAD_TIME': 'MERGE_TIME - FIRST_COMMIT_TIME',
                'LEAD_TIME_HOURS': '(MERGE_TIME - FIRST_COMMIT_TIME) / 3600000',
                'LEAD_TIME_DAYS': '(MERGE_TIME - FIRST_COMMIT_TIME) / 86400000',
                'MEAN': 'SUM(LEAD_TIME_VALUES) / COUNT(LEAD_TIME_VALUES)',
                'MEDIAN': 'MIDDLE_VALUE(SORTED_LEAD_TIME_VALUES)',
                'P75': 'VALUE_AT_75TH_PERCENTILE(SORTED_LEAD_TIME_VALUES)',
                'P95': 'VALUE_AT_95TH_PERCENTILE(SORTED_LEAD_TIME_VALUES)'
            }
        
        # Formula descriptions
        descriptions = {
            'LEAD_TIME': 'Time difference between first commit and merge',
            'LEAD_TIME_HOURS': 'Lead time converted to hours',
            'LEAD_TIME_DAYS': 'Lead time converted to days',
            'MEAN': 'Average lead time across all PRs',
            'MEDIAN': 'Middle value when lead times are sorted',
            'P75': '75th percentile of lead times',
            'P95': '95th percentile of lead times',
            'CONTRIBUTOR_EFFICIENCY': 'Individual contributor merge success rate',
            'AVG_CONTRIBUTOR_LEAD_TIME': 'Average lead time per contributor'
        }
        
        # Examples
        examples = {
            'LEAD_TIME': '2024-01-15 10:00 - 2024-01-10 14:00 = 4.83 days',
            'LEAD_TIME_HOURS': '4.83 days = 115.9 hours',
            'LEAD_TIME_DAYS': '115.9 hours = 4.83 days',
            'MEAN': '(2.5 + 4.0 + 1.2 + 6.8) / 4 = 3.625 days',
            'MEDIAN': 'Sorted: [1.2, 2.5, 4.0, 6.8] → Median = 3.25 days',
            'P75': '75% of PRs complete within this time',
            'P95': '95% of PRs complete within this time'
        }
        
        # Add formula data
        row = 4
        for formula_name, formula_text in formulas.items():
            ws[f'A{row}'] = formula_name
            ws[f'B{row}'] = formula_text
            ws[f'C{row}'] = descriptions.get(formula_name, 'Statistical calculation')
            ws[f'D{row}'] = examples.get(formula_name, 'See documentation')
            
            # Format cells
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = self.styles['border']
                if col == 2:  # Formula column
                    cell.font = Font(family='Consolas', size=10)
                if col in [3, 4]:  # Description and example columns
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 35
        
        # Set row height for better readability
        for row_num in range(4, row):
            ws.row_dimensions[row_num].height = 30
        
        return ws
    
    def generate_excel_report(self, output_path: str):
        """Generate the complete Excel report"""
        try:
            # Create all sheets
            logger.info("Creating Excel sheets...")
            
            # Create hidden data sheet first
            self.create_summary_data_sheet()
            
            # Create visible sheets
            self.create_summary_sheet()
            self.create_detailed_data_sheet()
            self.create_trends_sheet()
            self.create_contributors_sheet()
            
            # Create ML insights sheet if data is available
            if self.insights_data:
                self.create_ml_insights_sheet()
            
            # Create formulas documentation
            self.create_formulas_sheet()
            
            # Save workbook
            self.workbook.save(output_path)
            logger.info(f"Excel report saved successfully: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            raise

def main():
    parser = argparse.ArgumentParser(description='Generate Excel Report for Commit-to-Merge Lead Time Analysis')
    parser.add_argument('--input', '-i', required=True, help='Input JSON file path')
    parser.add_argument('--output', '-o', required=True, help='Output Excel file path')
    parser.add_argument('--insights', help='ML insights JSON file path (optional)')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose output')
    
    args = parser.parse_args()
    
    try:
        # Validate input file
        if not Path(args.input).exists():
            print(f"❌ Error: Input file not found: {args.input}")
            sys.exit(1)
        
        # Create output directory if needed
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        print("📊 Generating Excel report...")
        
        # Initialize generator
        generator = CommitMergeLeadTimeExcelGenerator()
        
        # Load data
        print(f"📂 Loading data from {args.input}...")
        generator.load_data(args.input, args.insights)
        
        # Generate report
        print("📈 Creating Excel sheets and charts...")
        output_file = generator.generate_excel_report(args.output)
        
        print("\n✅ EXCEL REPORT GENERATED SUCCESSFULLY!")
        print("=" * 50)
        print(f"📊 File saved: {output_file}")
        print(f"📋 Sheets created:")
        print("   • Executive Summary")
        print("   • Pull Request Details")
        print("   • Trends Analysis")
        print("   • Contributors Analysis")
        if generator.insights_data:
            print("   • ML Insights")
        print("   • Formulas Documentation")
        
        print("\n💡 Excel Features:")
        print("   • Interactive charts and visualizations")
        print("   • Conditional formatting for quick insights")
        print("   • Formula-based calculations with cell references")
        print("   • Multiple analysis perspectives")
        print("   • Professional formatting and styling")
        
    except Exception as e:
        print(f"❌ Failed to generate Excel report: {str(e)}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()